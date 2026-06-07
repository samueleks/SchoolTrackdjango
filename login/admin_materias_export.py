"""Generación de catálogo PDF/Excel de materias (administrativo)."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
from fpdf import FPDF
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .admin_views import (
    _EXCEL_COL_CONTENIDO,
    _PDF_COLOR_BORDE,
    _PDF_COLOR_TEXTO_SECUNDARIO,
    _fpdf_output_bytes,
    _pdf_dibujar_encabezados_tabla,
    _pdf_dibujar_fila_tabla,
    _pdf_fecha_legible,
    _pdf_insertar_logo,
    _pdf_texto_seguro,
    _excel_aplicar_margenes_hoja,
    _excel_celda_datos,
    _excel_insertar_logo,
)
from .alumno_boleta_export import (
    _BoletaAlumnoPDF,
    _pdf_alineaciones,
    _pdf_ancho_util,
    _pdf_layout_tabla,
)

_TITULO = 'Catálogo de Materias'

_CONFIG_COLUMNAS = {
    'Código': {'peso': 0.9, 'min': 22, 'max': 32, 'align': 'L'},
    'Materia': {'peso': 3.0, 'min': 50, 'max': 120, 'align': 'L'},
    'Semestre': {'peso': 0.7, 'min': 18, 'max': 24, 'align': 'C'},
    'Créditos': {'peso': 0.7, 'min': 18, 'max': 24, 'align': 'C'},
    'Estado': {'peso': 0.8, 'min': 20, 'max': 28, 'align': 'C'},
}

_HEADERS = list(_CONFIG_COLUMNAS.keys())

_ANCHO_EXCEL = {
    'Código': (14, 18),
    'Materia': (28, 52),
    'Semestre': (10, 14),
    'Créditos': (10, 14),
    'Estado': (12, 16),
}

_ALINEACION_EXCEL = {
    'Código': 'left',
    'Materia': 'left',
    'Semestre': 'center',
    'Créditos': 'center',
    'Estado': 'center',
}

_EXCEL_FILA_CONTENIDO = 3
_EXCEL_ANCHO_COL_LOGO = 12
_EXCEL_LOGO_PX = 72


def _pdf_dibujar_cabecera_materias(
    pdf: FPDF,
    *,
    fecha: datetime,
    total_registros: int,
    generado_por: str,
) -> None:
    y_inicio = pdf.t_margin
    tiene_logo = _pdf_insertar_logo(pdf, x=pdf.l_margin, y=y_inicio, ancho=18, alto=18)
    bloque_logo = 22
    x_texto = pdf.l_margin + (bloque_logo if tiene_logo else 0)
    ancho_texto = _pdf_ancho_util(pdf) - (bloque_logo if tiene_logo else 0)

    pdf.set_xy(x_texto, y_inicio)
    pdf.set_font('Arial', 'B', 18)
    pdf.set_text_color(17, 24, 39)
    pdf.cell(ancho_texto, 8, _pdf_texto_seguro(_TITULO), 0, 1, 'L')

    pdf.set_x(x_texto)
    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(*_PDF_COLOR_TEXTO_SECUNDARIO)
    pdf.cell(ancho_texto, 5, _pdf_texto_seguro('SchoolTrack · Sistema de gestión escolar'), 0, 1, 'L')

    if generado_por:
        pdf.set_x(x_texto)
        pdf.set_font('Arial', '', 9)
        pdf.cell(ancho_texto, 5, _pdf_texto_seguro(f'Administrativo: {generado_por}'), 0, 1, 'L')

    pdf.set_x(x_texto)
    pdf.set_font('Arial', '', 9)
    pdf.cell(
        ancho_texto,
        5,
        _pdf_texto_seguro(
            f'Exportado el {_pdf_fecha_legible(fecha)} · {total_registros} materia(s)'
        ),
        0,
        1,
        'L',
    )

    y_fin = max(pdf.get_y(), y_inicio + bloque_logo)
    pdf.set_y(y_fin + 2)
    pdf.set_draw_color(*_PDF_COLOR_BORDE)
    pdf.set_line_width(0.3)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(4)


def _fila_materia(materia: dict) -> list[str]:
    return [
        str(materia.get('codigo') or '---'),
        str(materia.get('nombre') or '---'),
        str(materia.get('semestre') or '---'),
        str(materia.get('creditos') if materia.get('creditos') is not None else '---'),
        'Activa' if materia.get('activa') else 'Inactiva',
    ]


def generar_pdf_catalogo_materias(
    materias: list[dict],
    *,
    ahora: datetime,
    generado_por: str = '',
) -> bytes:
    pdf = _BoletaAlumnoPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.set_margins(15, 15, 15)
    pdf.add_page()

    filas = [_fila_materia(materia) for materia in materias]
    if not filas:
        filas = [['---', 'Sin materias registradas', '---', '---', '---']]

    x_inicio, anchos = _pdf_layout_tabla(pdf, _HEADERS, filas, _CONFIG_COLUMNAS)
    ancho_tabla = sum(anchos)
    activas = sum(1 for materia in materias if materia.get('activa'))
    inactivas = len(materias) - activas

    _pdf_dibujar_cabecera_materias(
        pdf,
        fecha=ahora,
        total_registros=len(materias),
        generado_por=generado_por,
    )

    pdf.set_font('Arial', 'B', 10)
    pdf.set_text_color(55, 65, 81)
    pdf.cell(0, 6, _pdf_texto_seguro('Materias registradas en el sistema'), 0, 1, 'L')
    pdf.ln(2)

    alineaciones = _pdf_alineaciones(_HEADERS, _CONFIG_COLUMNAS)
    _pdf_dibujar_encabezados_tabla(pdf, _HEADERS, anchos, x_inicio=x_inicio)
    for indice, fila in enumerate(filas):
        _pdf_dibujar_fila_tabla(
            pdf,
            fila,
            anchos,
            alineaciones=alineaciones,
            x_inicio=x_inicio,
            fill=indice % 2 == 1,
            headers=_HEADERS,
        )

    pdf.ln(4)
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(17, 24, 39)
    pdf.set_x(x_inicio)
    pdf.cell(
        ancho_tabla,
        8,
        _pdf_texto_seguro(f'Activas: {activas} · Inactivas: {inactivas}'),
        0,
        1,
        'R',
    )

    return _fpdf_output_bytes(pdf)


def _excel_dibujar_cabecera_materias(
    worksheet,
    *,
    fecha: datetime,
    total_registros: int,
    generado_por: str,
    num_columnas: int,
) -> int:
    from openpyxl.utils import get_column_letter

    _excel_aplicar_margenes_hoja(worksheet)

    fila = _EXCEL_FILA_CONTENIDO
    col_logo = get_column_letter(_EXCEL_COL_CONTENIDO)
    tiene_logo = _excel_insertar_logo(worksheet, ancla=f'{col_logo}{fila}', tamaño=_EXCEL_LOGO_PX)
    col_texto_idx = _EXCEL_COL_CONTENIDO + (1 if tiene_logo else 0)
    col_texto = get_column_letter(col_texto_idx)
    end_col_letter = get_column_letter(_excel_celda_datos(num_columnas))

    title_font = Font(name='Arial', size=18, bold=True, color='111827')
    subtitle_font = Font(name='Arial', size=10, color='6B7280')
    meta_font = Font(name='Arial', size=9, color='6B7280')
    separador_border = Border(bottom=Side(style='thin', color='D1D5DB'))

    if tiene_logo:
        worksheet.column_dimensions[col_logo].width = _EXCEL_ANCHO_COL_LOGO
        for offset in range(4):
            worksheet.row_dimensions[fila + offset].height = 18

    for offset in range(4):
        worksheet.merge_cells(f'{col_texto}{fila + offset}:{end_col_letter}{fila + offset}')

    worksheet[f'{col_texto}{fila}'] = _TITULO
    worksheet[f'{col_texto}{fila}'].font = title_font
    worksheet[f'{col_texto}{fila}'].alignment = Alignment(horizontal='left', vertical='center')

    worksheet[f'{col_texto}{fila + 1}'] = 'SchoolTrack · Sistema de gestión escolar'
    worksheet[f'{col_texto}{fila + 1}'].font = subtitle_font
    worksheet[f'{col_texto}{fila + 1}'].alignment = Alignment(horizontal='left', vertical='center')

    if generado_por:
        worksheet[f'{col_texto}{fila + 2}'] = f'Administrativo: {generado_por}'
    else:
        worksheet[f'{col_texto}{fila + 2}'] = 'Catálogo administrativo de materias'
    worksheet[f'{col_texto}{fila + 2}'].font = meta_font
    worksheet[f'{col_texto}{fila + 2}'].alignment = Alignment(horizontal='left', vertical='center')

    worksheet[f'{col_texto}{fila + 3}'] = (
        f'Exportado el {_pdf_fecha_legible(fecha)} · {total_registros} materia(s)'
    )
    worksheet[f'{col_texto}{fila + 3}'].font = meta_font
    worksheet[f'{col_texto}{fila + 3}'].alignment = Alignment(horizontal='left', vertical='center')

    fila_separador = fila + 4
    for col_num in range(_EXCEL_COL_CONTENIDO, _excel_celda_datos(num_columnas) + 1):
        worksheet.cell(row=fila_separador, column=col_num).border = separador_border

    return fila_separador + 2


def generar_excel_catalogo_materias(
    materias: list[dict],
    *,
    ahora: datetime,
    generado_por: str = '',
) -> BytesIO:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.header_footer import HeaderFooterItem, _HeaderFooterPart

    headers = _HEADERS
    filas = [_fila_materia(materia) for materia in materias]
    activas = sum(1 for materia in materias if materia.get('activa'))
    inactivas = len(materias) - activas

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        worksheet = writer.book.create_sheet('Materias')
        writer.sheets['Materias'] = worksheet
        if 'Sheet1' in writer.book.sheetnames:
            del writer.book['Sheet1']

        fila_encabezados = _excel_dibujar_cabecera_materias(
            worksheet,
            fecha=ahora,
            total_registros=len(materias),
            generado_por=generado_por,
            num_columnas=len(headers),
        )

        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=9, color='1F2937')
        footer_font = Font(name='Arial', size=8, italic=True, color='9CA3AF')
        resumen_font = Font(name='Arial', size=9, bold=True, color='1F2937')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        border_color = 'D1D5DB'
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color),
        )

        for col_num, column in enumerate(headers, 1):
            col_excel = _excel_celda_datos(col_num)
            cell = worksheet.cell(row=fila_encabezados, column=col_excel, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for row_offset, row in enumerate(filas):
            fila_excel = fila_encabezados + 1 + row_offset
            es_zebra = row_offset % 2 == 1
            for col_num, value in enumerate(row, 1):
                col_excel = _excel_celda_datos(col_num)
                columna = headers[col_num - 1]
                cell = worksheet.cell(row=fila_excel, column=col_excel, value=value)
                cell.font = data_font
                cell.alignment = Alignment(
                    horizontal=_ALINEACION_EXCEL.get(columna, 'center'),
                    vertical='center',
                    wrap_text=columna == 'Materia',
                )
                cell.border = thin_border
                if es_zebra:
                    cell.fill = zebra_fill

        for col_num, column in enumerate(headers, 1):
            minimo, maximo = _ANCHO_EXCEL.get(column, (12, 30))
            largo_maximo = len(column)
            for fila in filas:
                if col_num - 1 < len(fila):
                    largo_maximo = max(largo_maximo, len(str(fila[col_num - 1])))
            worksheet.column_dimensions[get_column_letter(_excel_celda_datos(col_num))].width = max(
                minimo,
                min(maximo, largo_maximo + 2),
            )

        ultima_fila = fila_encabezados + len(filas) + 2
        col_pie = get_column_letter(_EXCEL_COL_CONTENIDO)
        end_col = get_column_letter(_excel_celda_datos(len(headers)))
        worksheet.merge_cells(f'{col_pie}{ultima_fila}:{end_col}{ultima_fila}')
        worksheet[f'{col_pie}{ultima_fila}'] = f'Activas: {activas} · Inactivas: {inactivas}'
        worksheet[f'{col_pie}{ultima_fila}'].font = resumen_font
        worksheet[f'{col_pie}{ultima_fila}'].alignment = Alignment(horizontal='right', vertical='center')

        ultima_fila += 1
        worksheet[f'{col_pie}{ultima_fila}'] = 'Generado automáticamente por SchoolTrack'
        worksheet[f'{col_pie}{ultima_fila}'].font = footer_font
        worksheet[f'{col_pie}{ultima_fila}'].alignment = Alignment(horizontal='left', vertical='center')

        worksheet.print_title_rows = f'{fila_encabezados}:{fila_encabezados}'
        worksheet.page_setup.orientation = 'portrait'
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.6
        worksheet.page_margins.right = 0.6
        worksheet.page_margins.top = 0.6
        worksheet.page_margins.bottom = 0.6
        pie_pagina = HeaderFooterItem()
        pie_pagina.left = _HeaderFooterPart('Generado automáticamente por SchoolTrack')
        pie_pagina.center = _HeaderFooterPart('Documento informativo · sin validez oficial')
        pie_pagina.right = _HeaderFooterPart('Página &P')
        worksheet.oddFooter = pie_pagina

    output.seek(0)
    return output
