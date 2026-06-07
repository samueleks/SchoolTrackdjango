"""Generación de reportes PDF/Excel para Consultar Reportes (maestro)."""
from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO

import pandas as pd
from fpdf import FPDF
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .admin_views import (
    _EXCEL_COL_CONTENIDO,
    _PDF_COLOR_BORDE,
    _PDF_COLOR_ENCABEZADO,
    _PDF_COLOR_NA,
    _PDF_COLOR_TEXTO_SECUNDARIO,
    _PDF_COLOR_ZEBRA,
    _PDF_VALOR_NO_APLICA,
    _excel_aplicar_margenes_hoja,
    _excel_celda_datos,
    _excel_insertar_logo,
    _fpdf_output_bytes,
    _pdf_dibujar_encabezados_tabla,
    _pdf_dibujar_fila_tabla,
    _pdf_es_celda_vacia,
    _pdf_fecha_legible,
    _pdf_insertar_logo,
    _pdf_rect_borde,
    _pdf_ruta_logo,
    _pdf_texto_seguro,
)

logger = logging.getLogger(__name__)

_TITULO_ASISTENCIAS = 'Reporte de Asistencias'
_TITULO_CALIFICACIONES = 'Reporte de Calificaciones'

_CONFIG_COLUMNAS_ASISTENCIAS = {
    'Fecha': {'peso': 0.85, 'min': 20, 'max': 26, 'align': 'C'},
    'Unidad': {'peso': 0.45, 'min': 14, 'max': 18, 'align': 'C'},
    'Matrícula': {'peso': 0.9, 'min': 22, 'max': 30, 'align': 'L'},
    'Alumno': {'peso': 2.0, 'min': 28, 'max': 48, 'align': 'L'},
    'Materia': {'peso': 1.8, 'min': 30, 'max': 70, 'align': 'L'},
    'Grupo': {'peso': 0.55, 'min': 16, 'max': 22, 'align': 'C'},
    'Ciclo': {'peso': 1.1, 'min': 24, 'max': 40, 'align': 'L'},
    'Horario': {'peso': 1.2, 'min': 26, 'max': 42, 'align': 'C'},
    'Estatus': {'peso': 0.85, 'min': 20, 'max': 28, 'align': 'C'},
    'Observaciones': {'peso': 2.4, 'min': 28, 'max': 90, 'align': 'L'},
}

_CONFIG_COLUMNAS_CALIFICACIONES = {
    'Fecha': {'peso': 0.85, 'min': 20, 'max': 26, 'align': 'C'},
    'Unidad': {'peso': 0.45, 'min': 14, 'max': 18, 'align': 'C'},
    'Matrícula': {'peso': 0.9, 'min': 22, 'max': 30, 'align': 'L'},
    'Alumno': {'peso': 2.0, 'min': 28, 'max': 48, 'align': 'L'},
    'Materia': {'peso': 1.8, 'min': 30, 'max': 70, 'align': 'L'},
    'Grupo': {'peso': 0.55, 'min': 16, 'max': 22, 'align': 'C'},
    'Ciclo': {'peso': 1.1, 'min': 24, 'max': 40, 'align': 'L'},
    'Calificación': {'peso': 0.75, 'min': 18, 'max': 24, 'align': 'C'},
    'Observaciones': {'peso': 2.4, 'min': 28, 'max': 90, 'align': 'L'},
}

_ANCHO_EXCEL_REPORTES = {
    'Fecha': (12, 14),
    'Unidad': (8, 10),
    'Matrícula': (14, 18),
    'Alumno': (22, 36),
    'Materia': (24, 48),
    'Grupo': (10, 14),
    'Ciclo': (18, 32),
    'Horario': (18, 30),
    'Estatus': (14, 20),
    'Calificación': (12, 16),
    'Observaciones': (20, 50),
}

_COLUMNAS_WRAP_EXCEL = frozenset({'Alumno', 'Materia', 'Ciclo', 'Observaciones'})

_TITULO_SESION_ASISTENCIAS = 'Pase de lista por sesión'

_CONFIG_COLUMNAS_SESION = {
    '#': {'peso': 0.35, 'min': 10, 'max': 14, 'align': 'C'},
    'Matrícula': {'peso': 0.9, 'min': 22, 'max': 30, 'align': 'L'},
    'Alumno': {'peso': 2.5, 'min': 30, 'max': 55, 'align': 'L'},
    'Estatus': {'peso': 0.9, 'min': 20, 'max': 28, 'align': 'C'},
    'Observaciones': {'peso': 2.2, 'min': 28, 'max': 90, 'align': 'L'},
}

_ANCHO_EXCEL_SESION = {
    '#': (6, 8),
    'Matrícula': (14, 16),
    'Alumno': (30, 42),
    'Estatus': (14, 18),
    'Observaciones': (16, 24),
}

_ANCHO_CABECERA_SESION = len(_CONFIG_COLUMNAS_ASISTENCIAS)
_EXCEL_ANCHO_COL_LOGO = 12
_EXCEL_LOGO_PX = 72

_HEADERS_SESION = list(_CONFIG_COLUMNAS_SESION.keys())
_COLUMNAS_WRAP_EXCEL_SESION = frozenset({'Alumno', 'Observaciones'})


class _ReporteMaestroPDF(FPDF):
    def footer(self):
        self.set_y(-14)
        self.set_draw_color(*_PDF_COLOR_BORDE)
        self.set_line_width(0.2)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(2)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(*_PDF_COLOR_TEXTO_SECUNDARIO)
        self.cell(0, 4, _pdf_texto_seguro('Generado automáticamente por SchoolTrack'), 0, 0, 'L')
        self.cell(0, 4, _pdf_texto_seguro(f'Página {self.page_no()}'), 0, 0, 'R')


def _pdf_ancho_util(pdf: FPDF) -> float:
    return pdf.w - pdf.l_margin - pdf.r_margin


def _pdf_ancho_texto(pdf: FPDF, texto: str, *, padding: float = 5) -> float:
    return pdf.get_string_width(_pdf_texto_seguro(texto)) + padding


def _headers_reporte(reporte_tipo: str) -> list[str]:
    if reporte_tipo == 'calificaciones':
        return list(_CONFIG_COLUMNAS_CALIFICACIONES.keys())
    return list(_CONFIG_COLUMNAS_ASISTENCIAS.keys())


def _config_columnas(reporte_tipo: str) -> dict:
    if reporte_tipo == 'calificaciones':
        return _CONFIG_COLUMNAS_CALIFICACIONES
    return _CONFIG_COLUMNAS_ASISTENCIAS


def _titulo_reporte(reporte_tipo: str) -> str:
    if reporte_tipo == 'calificaciones':
        return _TITULO_CALIFICACIONES
    return _TITULO_ASISTENCIAS


def _fila_exportacion(reporte: dict, reporte_tipo: str) -> list[str]:
    observaciones = reporte.get('observaciones') or ''
    base = [
        str(reporte.get('fecha', '')),
        str(reporte.get('unidad', '')),
        str(reporte.get('matricula', '')),
        str(reporte.get('alumno', '')),
        str(reporte.get('materia', '')),
        str(reporte.get('grupo', '')),
        str(reporte.get('ciclo', '')),
    ]
    if reporte_tipo == 'calificaciones':
        base.extend([
            str(reporte.get('calificacion', '')),
            observaciones,
        ])
    else:
        base.extend([
            str(reporte.get('horario', '')),
            str(reporte.get('estado', '')),
            observaciones,
        ])
    return base


def _pdf_ancho_minimo_columna(
    pdf: FPDF,
    header: str,
    col_idx: int,
    filas: list[list[str]],
    config: dict,
) -> float:
    pdf.set_font('Arial', 'B', 9)
    ancho = _pdf_ancho_texto(pdf, header, padding=6)
    pdf.set_font('Arial', '', 8)
    ancho = max(ancho, _pdf_ancho_texto(pdf, _PDF_VALOR_NO_APLICA))
    for fila in filas:
        if col_idx >= len(fila) or _pdf_es_celda_vacia(fila[col_idx]):
            continue
        ancho = max(ancho, _pdf_ancho_texto(pdf, fila[col_idx]))
    cfg = config.get(header, {})
    minimo = cfg.get('min', 18)
    maximo = cfg.get('max', 55)
    return min(maximo, max(minimo, ancho))


def _pdf_layout_tabla(
    pdf: FPDF,
    headers: list[str],
    filas: list[list[str]],
    config: dict,
) -> tuple[float, list[float]]:
    anchos = [
        _pdf_ancho_minimo_columna(pdf, header, col_idx, filas, config)
        for col_idx, header in enumerate(headers)
    ]
    ancho_util = _pdf_ancho_util(pdf)
    total = sum(anchos)

    if total > ancho_util:
        factor = ancho_util / total
        anchos = [ancho * factor for ancho in anchos]
    else:
        sobrante = ancho_util - total
        while sobrante > 0.05:
            indices = [
                i for i, header in enumerate(headers)
                if anchos[i] < config.get(header, {}).get('max', 90)
            ]
            if not indices:
                break
            pesos = [config.get(headers[i], {}).get('peso', 1.0) for i in indices]
            suma_pesos = sum(pesos) or 1
            asignado = 0.0
            for indice, peso in zip(indices, pesos):
                tope = config.get(headers[indice], {}).get('max', 90)
                incremento = min(tope - anchos[indice], sobrante * (peso / suma_pesos))
                anchos[indice] += incremento
                asignado += incremento
            if asignado <= 0.05:
                break
            sobrante -= asignado

    return pdf.l_margin, anchos


def _pdf_alineaciones(headers: list[str], config: dict) -> list[str]:
    return [config.get(header, {}).get('align', 'C') for header in headers]


def _pdf_dibujar_cabecera_documento(
    pdf: FPDF,
    *,
    titulo: str,
    fecha: datetime,
    total_registros: int,
    maestro_nombre: str,
    filtros: list[str] | None = None,
    filtros_por_fila: bool = False,
    responsable_etiqueta: str = 'Maestro',
) -> None:
    y_inicio = pdf.t_margin
    tiene_logo = _pdf_insertar_logo(pdf, x=pdf.l_margin, y=y_inicio, ancho=18, alto=18)
    bloque_logo = 22
    x_texto = pdf.l_margin + (bloque_logo if tiene_logo else 0)
    ancho_texto = _pdf_ancho_util(pdf) - (bloque_logo if tiene_logo else 0)

    pdf.set_xy(x_texto, y_inicio)
    pdf.set_font('Arial', 'B', 18)
    pdf.set_text_color(17, 24, 39)
    pdf.cell(ancho_texto, 8, _pdf_texto_seguro(titulo), 0, 1, 'L')

    pdf.set_x(x_texto)
    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(*_PDF_COLOR_TEXTO_SECUNDARIO)
    pdf.cell(ancho_texto, 5, _pdf_texto_seguro('SchoolTrack · Sistema de gestión escolar'), 0, 1, 'L')

    pdf.set_x(x_texto)
    pdf.set_font('Arial', '', 9)
    pdf.cell(
        ancho_texto,
        5,
        _pdf_texto_seguro(f'{responsable_etiqueta}: {maestro_nombre}'),
        0,
        1,
        'L',
    )

    pdf.set_x(x_texto)
    pdf.cell(
        ancho_texto,
        5,
        _pdf_texto_seguro(
            f'Exportado el {_pdf_fecha_legible(fecha)} · {total_registros} registro(s)'
        ),
        0,
        1,
        'L',
    )

    if filtros:
        pdf.set_font('Arial', 'I', 8)
        if filtros_por_fila:
            pdf.set_x(x_texto)
            pdf.cell(ancho_texto, 5, _pdf_texto_seguro('Detalle de la sesión'), 0, 1, 'L')
            for filtro in filtros:
                pdf.set_x(x_texto)
                pdf.cell(ancho_texto, 5, _pdf_texto_seguro(f'  · {filtro}'), 0, 1, 'L')
        else:
            pdf.set_x(x_texto)
            pdf.multi_cell(
                ancho_texto,
                4,
                _pdf_texto_seguro('Filtros: ' + ' · '.join(filtros)),
                0,
                'L',
            )

    y_fin = max(pdf.get_y(), y_inicio + bloque_logo)
    pdf.set_y(y_fin + 2)
    pdf.set_draw_color(*_PDF_COLOR_BORDE)
    pdf.set_line_width(0.3)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(4)


def _excel_dibujar_cabecera_documento(
    worksheet,
    *,
    titulo: str,
    num_columnas: int,
    fecha: datetime,
    total_registros: int,
    maestro_nombre: str,
    filtros: list[str] | None = None,
    ancho_merge_cabecera: int | None = None,
    filtros_por_fila: bool = False,
    logo_compacto: bool = False,
    responsable_etiqueta: str = 'Maestro',
) -> int:
    from openpyxl.utils import get_column_letter

    _excel_aplicar_margenes_hoja(worksheet)

    fila = 3
    col_logo = get_column_letter(_EXCEL_COL_CONTENIDO)
    ancla_logo = f'{col_logo}{fila}'
    logo_px = 48 if logo_compacto else _EXCEL_LOGO_PX
    tiene_logo = _excel_insertar_logo(worksheet, ancla=ancla_logo, tamaño=logo_px)
    col_texto_idx = _EXCEL_COL_CONTENIDO + (1 if tiene_logo else 0)
    col_texto = get_column_letter(col_texto_idx)
    merge_columnas = max(ancho_merge_cabecera or num_columnas, num_columnas)
    end_col_letter = get_column_letter(_excel_celda_datos(merge_columnas))
    if merge_columnas > num_columnas:
        for col_idx in range(num_columnas + 1, merge_columnas + 1):
            worksheet.column_dimensions[get_column_letter(_excel_celda_datos(col_idx))].width = 12
    title_font = Font(name='Arial', size=18, bold=True, color='111827')
    subtitle_font = Font(name='Arial', size=10, color='6B7280')
    meta_font = Font(name='Arial', size=9, color='6B7280')
    filtros_font = Font(name='Arial', size=8, italic=True, color='6B7280')
    filtros_titulo_font = Font(name='Arial', size=8, bold=True, color='6B7280')
    separador_border = Border(bottom=Side(style='thin', color='D1D5DB'))

    if tiene_logo:
        worksheet.column_dimensions[col_logo].width = (
            10 if logo_compacto else _EXCEL_ANCHO_COL_LOGO
        )
        altura_fila_logo = max(16, logo_px // 4)
        for offset in range(4):
            worksheet.row_dimensions[fila + offset].height = altura_fila_logo

    for offset in range(4):
        worksheet.merge_cells(f'{col_texto}{fila + offset}:{end_col_letter}{fila + offset}')

    worksheet[f'{col_texto}{fila}'] = titulo
    worksheet[f'{col_texto}{fila}'].font = title_font
    worksheet[f'{col_texto}{fila}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    worksheet[f'{col_texto}{fila + 1}'] = 'SchoolTrack · Sistema de gestión escolar'
    worksheet[f'{col_texto}{fila + 1}'].font = subtitle_font
    worksheet[f'{col_texto}{fila + 1}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    worksheet[f'{col_texto}{fila + 2}'] = f'{responsable_etiqueta}: {maestro_nombre}'
    worksheet[f'{col_texto}{fila + 2}'].font = meta_font
    worksheet[f'{col_texto}{fila + 2}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    worksheet[f'{col_texto}{fila + 3}'] = (
        f'Exportado el {_pdf_fecha_legible(fecha)} · {total_registros} registro(s)'
    )
    worksheet[f'{col_texto}{fila + 3}'].font = meta_font
    worksheet[f'{col_texto}{fila + 3}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    fila_separador = fila + 4
    if filtros:
        if filtros_por_fila:
            worksheet[f'{col_texto}{fila_separador}'] = 'Detalle de la sesión'
            worksheet[f'{col_texto}{fila_separador}'].font = filtros_titulo_font
            worksheet.merge_cells(f'{col_texto}{fila_separador}:{end_col_letter}{fila_separador}')
            worksheet[f'{col_texto}{fila_separador}'].alignment = Alignment(horizontal='left', vertical='center')
            fila_separador += 1
            for filtro in filtros:
                worksheet[f'{col_texto}{fila_separador}'] = filtro
                worksheet[f'{col_texto}{fila_separador}'].font = filtros_font
                worksheet.merge_cells(f'{col_texto}{fila_separador}:{end_col_letter}{fila_separador}')
                worksheet[f'{col_texto}{fila_separador}'].alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True,
                )
                worksheet.row_dimensions[fila_separador].height = 18
                fila_separador += 1
        else:
            worksheet[f'{col_texto}{fila_separador}'] = 'Filtros: ' + ' · '.join(filtros)
            worksheet[f'{col_texto}{fila_separador}'].font = filtros_font
            worksheet.merge_cells(f'{col_texto}{fila_separador}:{end_col_letter}{fila_separador}')
            worksheet[f'{col_texto}{fila_separador}'].alignment = Alignment(
                horizontal='left',
                vertical='center',
                wrap_text=False,
                shrink_to_fit=True,
            )
            worksheet.row_dimensions[fila_separador].height = 18
            fila_separador += 1

    for col_num in range(_EXCEL_COL_CONTENIDO, _excel_celda_datos(merge_columnas) + 1):
        cell = worksheet.cell(row=fila_separador, column=col_num)
        cell.border = separador_border

    return fila_separador + 2


def _excel_aplicar_anchos_columnas(
    worksheet,
    headers: list[str],
    filas: list,
    anchos: dict,
    *,
    reservar_col_logo: bool = False,
) -> None:
    from openpyxl.utils import get_column_letter

    for col_num, column in enumerate(headers, 1):
        minimo, maximo = anchos.get(column, (12, 30))
        largo_maximo = len(column)
        for fila in filas:
            if col_num - 1 < len(fila):
                largo_maximo = max(largo_maximo, len(str(fila[col_num - 1])))
        ancho = max(minimo, min(maximo, largo_maximo + 2))
        if reservar_col_logo and col_num == 1:
            ancho = max(ancho, _EXCEL_ANCHO_COL_LOGO)
        worksheet.column_dimensions[get_column_letter(_excel_celda_datos(col_num))].width = ancho


def _excel_alineacion_celda(valor: str, columna: str, config: dict) -> str:
    if _pdf_es_celda_vacia(valor):
        return 'center'
    align = config.get(columna, {}).get('align', 'C')
    return {'L': 'left', 'C': 'center', 'R': 'right'}.get(align, 'center')


def generar_pdf_reportes_maestro(
    reportes: list[dict],
    *,
    reporte_tipo: str,
    maestro_nombre: str,
    filtros: list[str] | None,
    ahora: datetime,
    responsable_etiqueta: str = 'Maestro',
) -> bytes:
    headers = _headers_reporte(reporte_tipo)
    config = _config_columnas(reporte_tipo)
    filas = [_fila_exportacion(reporte, reporte_tipo) for reporte in reportes]

    pdf = _ReporteMaestroPDF(orientation='L', unit='mm', format='A4')
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    _pdf_dibujar_cabecera_documento(
        pdf,
        titulo=_titulo_reporte(reporte_tipo),
        fecha=ahora,
        total_registros=len(reportes),
        maestro_nombre=maestro_nombre,
        filtros=filtros,
        responsable_etiqueta=responsable_etiqueta,
    )

    x_tabla, col_widths = _pdf_layout_tabla(pdf, headers, filas, config)
    alineaciones = _pdf_alineaciones(headers, config)
    _pdf_dibujar_encabezados_tabla(pdf, headers, col_widths, x_inicio=x_tabla)

    for indice, fila in enumerate(filas):
        _pdf_dibujar_fila_tabla(
            pdf,
            fila,
            col_widths,
            alineaciones=alineaciones,
            x_inicio=x_tabla,
            fill=indice % 2 == 1,
            headers=headers,
        )

    return _fpdf_output_bytes(pdf)


def generar_excel_reportes_maestro(
    reportes: list[dict],
    *,
    reporte_tipo: str,
    maestro_nombre: str,
    filtros: list[str] | None,
    ahora: datetime,
    responsable_etiqueta: str = 'Maestro',
) -> BytesIO:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.header_footer import HeaderFooterItem, _HeaderFooterPart

    headers = _headers_reporte(reporte_tipo)
    config = _config_columnas(reporte_tipo)
    filas = [_fila_exportacion(reporte, reporte_tipo) for reporte in reportes]
    nombre_hoja = 'Asistencias' if reporte_tipo != 'calificaciones' else 'Calificaciones'

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        worksheet = writer.book.create_sheet(nombre_hoja)
        writer.sheets[nombre_hoja] = worksheet
        if 'Sheet1' in writer.book.sheetnames:
            del writer.book['Sheet1']

        fila_encabezados = _excel_dibujar_cabecera_documento(
            worksheet,
            titulo=_titulo_reporte(reporte_tipo),
            num_columnas=len(headers),
            fecha=ahora,
            total_registros=len(reportes),
            maestro_nombre=maestro_nombre,
            filtros=filtros,
            responsable_etiqueta=responsable_etiqueta,
        )

        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=9, color='1F2937')
        na_font = Font(name='Arial', size=9, color='BEC4D0')
        footer_font = Font(name='Arial', size=8, italic=True, color='9CA3AF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        border_color = 'D1D5DB'
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color),
        )

        for col_num, column in enumerate(headers, 1):
            col_excel = _excel_celda_datos(col_num)
            cell = worksheet.cell(row=fila_encabezados, column=col_excel, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for row_offset, row in enumerate(filas):
            fila_excel = fila_encabezados + 1 + row_offset
            es_zebra = row_offset % 2 == 1
            for col_num, value in enumerate(row, 1):
                col_excel = _excel_celda_datos(col_num)
                cell = worksheet.cell(row=fila_excel, column=col_excel)
                columna = headers[col_num - 1]
                texto = '' if pd.isna(value) else str(value)
                cell.value = texto
                cell.font = na_font if texto == _PDF_VALOR_NO_APLICA else data_font
                cell.alignment = Alignment(
                    horizontal=_excel_alineacion_celda(texto, columna, config),
                    vertical='center',
                    wrap_text=columna in _COLUMNAS_WRAP_EXCEL,
                )
                cell.border = thin_border
                if es_zebra:
                    cell.fill = zebra_fill

        _excel_aplicar_anchos_columnas(
            worksheet,
            headers,
            filas,
            _ANCHO_EXCEL_REPORTES,
            reservar_col_logo=True,
        )

        col_pie = get_column_letter(_EXCEL_COL_CONTENIDO)
        ultima_fila = fila_encabezados + len(filas) + 2
        worksheet[f'{col_pie}{ultima_fila}'] = 'Generado automáticamente por SchoolTrack'
        worksheet[f'{col_pie}{ultima_fila}'].font = footer_font
        worksheet[f'{col_pie}{ultima_fila}'].alignment = Alignment(horizontal='left', vertical='center')

        worksheet.print_title_rows = f'{fila_encabezados}:{fila_encabezados}'
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.6
        worksheet.page_margins.right = 0.6
        worksheet.page_margins.top = 0.6
        worksheet.page_margins.bottom = 0.6
        pie_pagina = HeaderFooterItem()
        pie_pagina.left = _HeaderFooterPart('Generado automáticamente por SchoolTrack')
        pie_pagina.right = _HeaderFooterPart('Página &P')
        worksheet.oddFooter = pie_pagina

    output.seek(0)
    return output


def _valor_excel_sesion(columna: str, valor) -> str | int:
    texto = '' if valor in (None, '') else str(valor).strip()
    if columna == '#' and texto.isdigit():
        return int(texto)
    if columna == 'Matrícula' and texto.isdigit():
        return int(texto)
    return texto


def _filas_sesion_exportacion(sesion_consulta: dict) -> list[list]:
    filas = []
    for indice, alumno in enumerate(sesion_consulta.get('alumnos', []), 1):
        filas.append([
            indice,
            alumno.get('matricula', ''),
            str(alumno.get('nombre', '')),
            str(alumno.get('estado', '')),
            str(alumno.get('observaciones') or ''),
        ])
    return filas


def generar_pdf_sesion_asistencia_maestro(
    sesion_consulta: dict,
    *,
    maestro_nombre: str,
    filtros: list[str] | None,
    ahora: datetime,
) -> bytes:
    headers = _HEADERS_SESION
    config = _CONFIG_COLUMNAS_SESION
    filas = [[str(valor) for valor in fila] for fila in _filas_sesion_exportacion(sesion_consulta)]

    pdf = _ReporteMaestroPDF(orientation='L', unit='mm', format='A4')
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    _pdf_dibujar_cabecera_documento(
        pdf,
        titulo=_TITULO_SESION_ASISTENCIAS,
        fecha=ahora,
        total_registros=len(filas),
        maestro_nombre=maestro_nombre,
        filtros=filtros,
    )

    x_tabla, col_widths = _pdf_layout_tabla(pdf, headers, filas, config)
    alineaciones = _pdf_alineaciones(headers, config)
    _pdf_dibujar_encabezados_tabla(pdf, headers, col_widths, x_inicio=x_tabla)

    for indice, fila in enumerate(filas):
        _pdf_dibujar_fila_tabla(
            pdf,
            fila,
            col_widths,
            alineaciones=alineaciones,
            x_inicio=x_tabla,
            fill=indice % 2 == 1,
            headers=headers,
        )

    return _fpdf_output_bytes(pdf)


def generar_excel_sesion_asistencia_maestro(
    sesion_consulta: dict,
    *,
    maestro_nombre: str,
    filtros: list[str] | None,
    ahora: datetime,
) -> BytesIO:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.header_footer import HeaderFooterItem, _HeaderFooterPart

    headers = _HEADERS_SESION
    config = _CONFIG_COLUMNAS_SESION
    filas = _filas_sesion_exportacion(sesion_consulta)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        worksheet = writer.book.create_sheet('Pase de lista')
        writer.sheets['Pase de lista'] = worksheet
        if 'Sheet1' in writer.book.sheetnames:
            del writer.book['Sheet1']

        fila_encabezados = _excel_dibujar_cabecera_documento(
            worksheet,
            titulo=_TITULO_SESION_ASISTENCIAS,
            num_columnas=len(headers),
            fecha=ahora,
            total_registros=len(filas),
            maestro_nombre=maestro_nombre,
            filtros=filtros,
            ancho_merge_cabecera=_ANCHO_CABECERA_SESION,
        )

        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=9, color='1F2937')
        na_font = Font(name='Arial', size=9, color='BEC4D0')
        footer_font = Font(name='Arial', size=8, italic=True, color='9CA3AF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        border_color = 'D1D5DB'
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color),
        )

        for col_num, column in enumerate(headers, 1):
            col_excel = _excel_celda_datos(col_num)
            cell = worksheet.cell(row=fila_encabezados, column=col_excel, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for row_offset, row in enumerate(filas):
            fila_excel = fila_encabezados + 1 + row_offset
            es_zebra = row_offset % 2 == 1
            for col_num, value in enumerate(row, 1):
                col_excel = _excel_celda_datos(col_num)
                cell = worksheet.cell(row=fila_excel, column=col_excel)
                columna = headers[col_num - 1]
                if pd.isna(value):
                    valor_celda = ''
                else:
                    valor_celda = _valor_excel_sesion(columna, value)
                cell.value = valor_celda
                texto = '' if valor_celda == '' else str(valor_celda)
                cell.font = na_font if texto in ('', _PDF_VALOR_NO_APLICA, 'Sin registro') else data_font
                cell.alignment = Alignment(
                    horizontal=_excel_alineacion_celda(texto, columna, config),
                    vertical='center',
                    wrap_text=columna in _COLUMNAS_WRAP_EXCEL_SESION,
                )
                cell.border = thin_border
                if es_zebra:
                    cell.fill = zebra_fill

        _excel_aplicar_anchos_columnas(
            worksheet,
            headers,
            filas,
            _ANCHO_EXCEL_SESION,
            reservar_col_logo=True,
        )

        col_pie = get_column_letter(_EXCEL_COL_CONTENIDO)
        ultima_fila = fila_encabezados + len(filas) + 2
        worksheet[f'{col_pie}{ultima_fila}'] = 'Generado automáticamente por SchoolTrack'
        worksheet[f'{col_pie}{ultima_fila}'].font = footer_font
        worksheet[f'{col_pie}{ultima_fila}'].alignment = Alignment(horizontal='left', vertical='center')

        worksheet.print_title_rows = f'{fila_encabezados}:{fila_encabezados}'
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.6
        worksheet.page_margins.right = 0.6
        worksheet.page_margins.top = 0.6
        worksheet.page_margins.bottom = 0.6
        pie_pagina = HeaderFooterItem()
        pie_pagina.left = _HeaderFooterPart('Generado automáticamente por SchoolTrack')
        pie_pagina.right = _HeaderFooterPart('Página &P')
        worksheet.oddFooter = pie_pagina

    output.seek(0)
    return output
