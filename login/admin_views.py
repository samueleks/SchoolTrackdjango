import os
import json
import logging
import shutil
from io import BytesIO
from datetime import datetime
import unicodedata
import re
from urllib.parse import urlencode
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db import connection, transaction
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.conf import settings
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF

from .models import Usuarios, Alumnos, Maestros, Administrativos, Administrador, DatosPersonales, Carrera, CicloEscolar, Grupo, LogCalificacion
from .password_utils import generar_contrasena_temporal


logger = logging.getLogger(__name__)


def _fpdf_output_bytes(pdf: FPDF) -> bytes:
    """fpdf2 devuelve str en versiones viejas y bytearray/bytes en versiones nuevas."""
    content = pdf.output(dest='S')
    if isinstance(content, str):
        return content.encode('latin-1')
    return bytes(content)


_PDF_VALOR_NO_APLICA = '-----'
_ROLES_PDF_USUARIOS = ('Alumno', 'Maestro', 'Administrativo', 'Administrador')
_PESOS_COLUMNAS_PDF_USUARIOS = {
    'Matrícula': 0.9,
    'Nombre Completo': 2.2,
    'Rol': 1.0,
    'Carrera': 3.4,
    'Semestre': 0.55,
    'Departamento': 2.5,
    'Cubículo': 0.65,
    'Nivel Prioridad': 0.6,
}
_ANCHO_MINIMO_PDF_USUARIOS = {
    'Matrícula': 22,
    'Nombre Completo': 30,
    'Rol': 27,
    'Carrera': 34,
    'Semestre': 21,
    'Departamento': 30,
    'Cubículo': 14,
    'Nivel Prioridad': 15,
}
_ANCHO_MAXIMO_PDF_USUARIOS = {
    'Matrícula': 28,
    'Nombre Completo': 38,
    'Rol': 30,
    'Carrera': 95,
    'Semestre': 21,
    'Departamento': 68,
    'Cubículo': 20,
    'Nivel Prioridad': 20,
}
_COLUMNAS_EXPANSIBLES_PDF = frozenset({'Carrera', 'Nombre Completo', 'Departamento'})
_ALINEACION_PDF_USUARIOS = {
    'Matrícula': 'L',
    'Nombre Completo': 'L',
    'Rol': 'C',
    'Carrera': 'L',
    'Semestre': 'C',
    'Departamento': 'L',
    'Cubículo': 'C',
    'Nivel Prioridad': 'C',
}
_PDF_COLOR_ENCABEZADO = (30, 58, 138)
_PDF_COLOR_BORDE = (209, 213, 219)
_PDF_COLOR_ZEBRA = (248, 250, 252)
_PDF_COLOR_TEXTO_SECUNDARIO = (107, 114, 128)
_PDF_COLOR_NA = (190, 196, 208)
_MESES_ES_PDF = (
    'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
)


class _ReporteUsuariosPDF(FPDF):
    def footer(self):
        self.set_y(-14)
        self.set_draw_color(*_PDF_COLOR_BORDE)
        self.set_line_width(0.2)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(2)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(*_PDF_COLOR_TEXTO_SECUNDARIO)
        self.cell(0, 4, _pdf_texto_seguro('Generado automáticamente por SchoolTrack'), 0, 0, 'L')
        self.cell(0, 4, _pdf_texto_seguro(f'Página {self.page_no()}'), 0, 0, 'R')


def _pdf_texto_seguro(valor) -> str:
    texto = '' if valor in (None, '') else str(valor)
    return texto.encode('latin-1', 'replace').decode('latin-1')


def _pdf_fecha_legible(fecha: datetime) -> str:
    return f'{fecha.day:02d} de {_MESES_ES_PDF[fecha.month - 1]} de {fecha.year}'


def _pdf_ruta_logo() -> str | None:
    candidatos = [
        settings.BASE_DIR / 'login' / 'static' / 'logo.png',
        settings.STATIC_ROOT / 'logo.png',
    ]
    for ruta in candidatos:
        if ruta.exists():
            return str(ruta)
    return None


def _pdf_insertar_logo(pdf: FPDF, *, x: float, y: float, ancho: float, alto: float) -> bool:
    logo_ruta = _pdf_ruta_logo()
    if not logo_ruta:
        return False
    try:
        with open(logo_ruta, 'rb') as archivo:
            firma = archivo.read(4)
        tipo = 'JPEG' if firma[:2] == b'\xff\xd8' else 'PNG'
        pdf.image(logo_ruta, x=x, y=y, w=ancho, h=alto, type=tipo)
        return True
    except Exception:
        logger.warning('No se pudo cargar el logo para el PDF de usuarios', exc_info=True)
        return False


def _pdf_valor_celda(valor, *, no_aplica: bool = False) -> str:
    if no_aplica:
        return _PDF_VALOR_NO_APLICA
    if valor in (None, '', '-'):
        return ''
    return str(valor)


def _pdf_es_celda_vacia(valor: str) -> bool:
    return valor in ('', _PDF_VALOR_NO_APLICA)


def _pdf_nombre_completo(usuario: dict) -> str:
    return f"{usuario.get('Nombre', '')} {usuario.get('Apellido', '')}".strip()


def _pdf_ancho_util(pdf: FPDF) -> float:
    return pdf.w - pdf.l_margin - pdf.r_margin


def _pdf_ancho_texto(pdf: FPDF, texto: str, *, padding: float = 5) -> float:
    return pdf.get_string_width(_pdf_texto_seguro(texto)) + padding


def _pdf_ancho_minimo_columna(
    pdf: FPDF,
    header: str,
    col_idx: int,
    filas: list[list[str]],
) -> float:
    pdf.set_font('Arial', 'B', 9)
    ancho = _pdf_ancho_texto(pdf, header, padding=6)
    pdf.set_font('Arial', '', 8)

    if header == 'Rol':
        for rol in _ROLES_PDF_USUARIOS:
            ancho = max(ancho, _pdf_ancho_texto(pdf, rol))
    else:
        ancho = max(ancho, _pdf_ancho_texto(pdf, _PDF_VALOR_NO_APLICA))
        for fila in filas:
            if col_idx >= len(fila) or _pdf_es_celda_vacia(fila[col_idx]):
                continue
            ancho = max(ancho, _pdf_ancho_texto(pdf, fila[col_idx]))

    minimo = _ANCHO_MINIMO_PDF_USUARIOS.get(header, 18)
    maximo = _ANCHO_MAXIMO_PDF_USUARIOS.get(header, 55)
    return min(maximo, max(minimo, ancho))


def _pdf_layout_tabla_fluida(
    pdf: FPDF,
    headers: list[str],
    filas: list[list[str]],
) -> tuple[float, list[float]]:
    """Distribuye el ancho de la hoja respetando mínimos por columna y sin aplastar Rol/Semestre."""
    anchos = [
        _pdf_ancho_minimo_columna(pdf, header, col_idx, filas)
        for col_idx, header in enumerate(headers)
    ]

    ancho_util = _pdf_ancho_util(pdf)
    total = sum(anchos)

    if total > ancho_util:
        factor = ancho_util / total
        anchos = [ancho * factor for ancho in anchos]
    else:
        sobrante = ancho_util - total
        while sobrante > 0.05:
            indices = [
                i for i, header in enumerate(headers)
                if anchos[i] < _ANCHO_MAXIMO_PDF_USUARIOS.get(header, 90)
            ]
            if not indices:
                break

            pesos = [_PESOS_COLUMNAS_PDF_USUARIOS.get(headers[i], 1.0) for i in indices]
            suma_pesos = sum(pesos) or 1
            asignado = 0.0
            for indice, peso in zip(indices, pesos):
                tope = _ANCHO_MAXIMO_PDF_USUARIOS.get(headers[indice], 90)
                incremento = min(tope - anchos[indice], sobrante * (peso / suma_pesos))
                anchos[indice] += incremento
                asignado += incremento

            if asignado <= 0.05:
                break
            sobrante -= asignado

    return pdf.l_margin, anchos


def _pdf_alineaciones_columnas(headers: list[str]) -> list[str]:
    return [_ALINEACION_PDF_USUARIOS.get(header, 'C') for header in headers]


def _pdf_alineacion_celda(valor: str, align_default: str) -> str:
    if _pdf_es_celda_vacia(valor):
        return 'C'
    return align_default


def _pdf_rect_borde(pdf: FPDF, x: float, y: float, ancho: float, alto: float) -> None:
    pdf.set_draw_color(*_PDF_COLOR_BORDE)
    pdf.set_line_width(0.2)
    pdf.rect(x, y, ancho, alto)


def _pdf_partir_celda(
    pdf: FPDF,
    texto: str,
    ancho: float,
    altura_linea: float,
    *,
    align: str = 'C',
) -> list[str]:
    lineas = pdf.multi_cell(
        ancho,
        altura_linea,
        _pdf_texto_seguro(texto),
        border=0,
        align=align,
        split_only=True,
    )
    return lineas or ['']


def _pdf_dibujar_cabecera_documento(
    pdf: FPDF,
    *,
    fecha: datetime,
    total_registros: int,
    filtros: list[str] | None = None,
) -> None:
    y_inicio = pdf.t_margin
    tiene_logo = _pdf_insertar_logo(pdf, x=pdf.l_margin, y=y_inicio, ancho=18, alto=18)
    bloque_logo = 22
    x_texto = pdf.l_margin + (bloque_logo if tiene_logo else 0)
    ancho_texto = _pdf_ancho_util(pdf) - (bloque_logo if tiene_logo else 0)

    pdf.set_xy(x_texto, y_inicio)
    pdf.set_font('Arial', 'B', 18)
    pdf.set_text_color(17, 24, 39)
    pdf.cell(ancho_texto, 8, _pdf_texto_seguro('Reporte de Usuarios'), 0, 1, 'L')

    pdf.set_x(x_texto)
    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(*_PDF_COLOR_TEXTO_SECUNDARIO)
    pdf.cell(ancho_texto, 5, _pdf_texto_seguro('SchoolTrack · Sistema de gestión escolar'), 0, 1, 'L')

    pdf.set_x(x_texto)
    pdf.set_font('Arial', '', 9)
    pdf.cell(
        ancho_texto,
        5,
        _pdf_texto_seguro(
            f'Exportado el {_pdf_fecha_legible(fecha)} · {total_registros} registro(s)'
        ),
        0,
        1,
        'L',
    )

    if filtros:
        pdf.set_x(x_texto)
        pdf.set_font('Arial', 'I', 8)
        pdf.cell(ancho_texto, 5, _pdf_texto_seguro('Filtros: ' + ' · '.join(filtros)), 0, 1, 'L')

    y_fin = max(pdf.get_y(), y_inicio + bloque_logo)
    pdf.set_y(y_fin + 2)
    pdf.set_draw_color(*_PDF_COLOR_BORDE)
    pdf.set_line_width(0.3)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(4)


def _pdf_dibujar_encabezados_tabla(
    pdf: FPDF,
    headers: list[str],
    anchos: list[float],
    *,
    x_inicio: float | None = None,
    altura: float = 8,
) -> None:
    pdf.set_fill_color(*_PDF_COLOR_ENCABEZADO)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Arial', 'B', 9)
    x = x_inicio if x_inicio is not None else pdf.l_margin
    y = pdf.get_y()
    for header, ancho in zip(headers, anchos):
        pdf.set_xy(x, y)
        pdf.cell(ancho, altura, _pdf_texto_seguro(header), 0, 0, 'C', True)
        _pdf_rect_borde(pdf, x, y, ancho, altura)
        x += ancho
    pdf.set_xy(x_inicio if x_inicio is not None else pdf.l_margin, y + altura)


def _pdf_dibujar_fila_tabla(
    pdf: FPDF,
    valores: list[str],
    anchos: list[float],
    *,
    alineaciones: list[str] | None = None,
    x_inicio: float | None = None,
    altura_linea: float = 4.5,
    fill: bool = False,
    headers: list[str] | None = None,
) -> None:
    if alineaciones is None:
        alineaciones = ['C'] * len(valores)

    x = x_inicio if x_inicio is not None else pdf.l_margin
    y = pdf.get_y()
    alineaciones_efectivas = [
        _pdf_alineacion_celda(valor, align)
        for valor, align in zip(valores, alineaciones)
    ]
    lineas_por_celda = [
        _pdf_partir_celda(pdf, valor, ancho, altura_linea, align=align)
        for valor, ancho, align in zip(valores, anchos, alineaciones_efectivas)
    ]
    max_lineas = max((len(lineas) for lineas in lineas_por_celda), default=1)
    alto_fila = altura_linea * max_lineas

    if y + alto_fila > pdf.page_break_trigger:
        pdf.add_page()
        if headers:
            _pdf_dibujar_encabezados_tabla(pdf, headers, anchos, x_inicio=x)
        y = pdf.get_y()

    color_fondo = _PDF_COLOR_ZEBRA if fill else (255, 255, 255)
    pdf.set_font('Arial', '', 8)

    for indice, (lineas, ancho, align) in enumerate(zip(lineas_por_celda, anchos, alineaciones_efectivas)):
        x_celda = x + sum(anchos[:indice])
        pdf.set_fill_color(*color_fondo)
        pdf.rect(x_celda, y, ancho, alto_fila, style='F')
        _pdf_rect_borde(pdf, x_celda, y, ancho, alto_fila)

        es_vacio = valores[indice] == _PDF_VALOR_NO_APLICA
        pdf.set_text_color(*_PDF_COLOR_NA if es_vacio else (31, 41, 55))
        for num_linea, linea in enumerate(lineas):
            pdf.set_xy(x_celda, y + (num_linea * altura_linea))
            pdf.cell(ancho, altura_linea, linea, 0, 0, align)

    pdf.set_text_color(31, 41, 55)
    pdf.set_xy(x, y + alto_fila)


def _pdf_fila_usuario_exportacion(usuario: dict, rol_filtro: str) -> list[str]:
    rol = usuario.get('Rol', '')
    fila = [
        _pdf_valor_celda(usuario.get('Matrícula')),
        _pdf_nombre_completo(usuario),
        _pdf_valor_celda(usuario.get('Rol')),
    ]

    if rol_filtro == 'alumno':
        fila.extend([
            _pdf_valor_celda(usuario.get('Carrera')),
            _pdf_valor_celda(usuario.get('Semestre')),
        ])
    elif rol_filtro == 'maestro':
        fila.extend([
            _pdf_valor_celda(usuario.get('Departamento')),
            _pdf_valor_celda(usuario.get('Cubículo')),
        ])
    elif rol_filtro == 'administrativo':
        fila.extend([_pdf_valor_celda(usuario.get('Departamento'))])
    elif rol_filtro == 'admin':
        fila.extend([_pdf_valor_celda(usuario.get('Nivel Prioridad'))])
    else:
        fila.extend([
            _pdf_valor_celda(usuario.get('Carrera'), no_aplica=rol != 'Alumno'),
            _pdf_valor_celda(usuario.get('Semestre'), no_aplica=rol != 'Alumno'),
            _pdf_valor_celda(
                usuario.get('Departamento'),
                no_aplica=rol not in ('Maestro', 'Administrativo'),
            ),
        ])

    return fila


def _pdf_headers_usuarios(rol_filtro: str) -> list[str]:
    headers = ['Matrícula', 'Nombre Completo', 'Rol']
    if rol_filtro == 'alumno':
        headers.extend(['Carrera', 'Semestre'])
    elif rol_filtro == 'maestro':
        headers.extend(['Departamento', 'Cubículo'])
    elif rol_filtro == 'administrativo':
        headers.extend(['Departamento'])
    elif rol_filtro == 'admin':
        headers.extend(['Nivel Prioridad'])
    else:
        headers.extend(['Carrera', 'Semestre', 'Departamento'])
    return headers


_ROL_MAP_EXPORTACION = {
    'alumno': 'Alumno',
    'maestro': 'Maestro',
    'administrativo': 'Administrativo',
    'admin': 'Administrador',
}


def _mensaje_credenciales_temporales(tipo: str, matricula: str, nombre: str, contrasena: str) -> str:
    """Formato estructurado para el modal de credenciales en GestionUsuarios.html."""
    return f'TEMP_CRED|{tipo}|{matricula}|{nombre}|{contrasena}'


_ORDEN_CAMPOS_AGREGAR_USUARIO = (
    'rol', 'nombre', 'apellido', 'foto',
    'periodo_ingreso', 'carrera_id', 'semestre', 'estatus',
    'departamento', 'cubiculo', 'grado_academico',
    'puesto', 'nivel_prioridad', 'id_ciclo_escolar',
    'correo_inst', 'telefono', 'curp', 'fecha_nacimiento', 'genero',
)

_ORDEN_CAMPOS_EDITAR_USUARIO = (
    'foto', 'nombre', 'apellido',
    'carrera_id', 'semestre', 'estatus',
    'departamento', 'cubiculo', 'grado_academico',
    'puesto', 'nivel_prioridad',
    'correo_inst', 'telefono', 'curp', 'fecha_nacimiento', 'genero',
)


def _primer_campo_error_por_orden(errores_campos: dict, orden: tuple) -> str | None:
    for campo in orden:
        if campo in errores_campos:
            return campo
    return next(iter(errores_campos), None)


def _primer_campo_error_agregar_usuario(errores_campos: dict) -> str | None:
    return _primer_campo_error_por_orden(errores_campos, _ORDEN_CAMPOS_AGREGAR_USUARIO)


def _primer_campo_error_editar_usuario(errores_campos: dict) -> str | None:
    return _primer_campo_error_por_orden(errores_campos, _ORDEN_CAMPOS_EDITAR_USUARIO)


def _proximo_id_usuario_preview() -> int:
    """
    ID que la BD asignará al próximo INSERT en usuarios.
    En PostgreSQL el contador real es la secuencia, no MAX(id)+1 (los huecos por
    usuarios eliminados no se reutilizan).
    """
    from django.db.models import Max

    if connection.vendor == 'postgresql':
        table = Usuarios._meta.db_table
        column = Usuarios._meta.pk.column
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT pg_get_serial_sequence(%s, %s)',
                [table, column],
            )
            sequence_name = cursor.fetchone()[0]
            if sequence_name:
                cursor.execute(f'SELECT last_value, is_called FROM {sequence_name}')
                last_value, is_called = cursor.fetchone()
                return int(last_value + (1 if is_called else 0))

    max_id = Usuarios.objects.aggregate(Max('id_usuario'))['id_usuario__max'] or 0
    return max_id + 1
_ALINEACION_EXCEL_USUARIOS = {
    'Matrícula': 'left',
    'Nombre Completo': 'left',
    'Rol': 'center',
    'Carrera': 'left',
    'Semestre': 'center',
    'Departamento': 'left',
    'Cubículo': 'center',
    'Nivel Prioridad': 'center',
}
_ANCHO_EXCEL_USUARIOS = {
    'Matrícula': (12, 18),
    'Nombre Completo': (22, 34),
    'Rol': (16, 22),
    'Carrera': (28, 52),
    'Semestre': (10, 12),
    'Departamento': (24, 42),
    'Cubículo': (10, 14),
    'Nivel Prioridad': (12, 16),
}


def _recolectar_usuarios_exportacion() -> list[dict]:
    usuarios_data = []

    for alumno in Alumnos.objects.select_related('id_usuario', 'id_carrera').all():
        usuarios_data.append({
            'Matrícula': alumno.id_usuario.matricula,
            'Nombre': alumno.id_usuario.nombre,
            'Apellido': alumno.id_usuario.apellido,
            'Rol': 'Alumno',
            'Carrera': str(alumno.id_carrera) if alumno.id_carrera else '',
            'Semestre': alumno.semestre,
            'Estatus': alumno.estatus,
        })

    for maestro in Maestros.objects.select_related('id_usuario').all():
        usuarios_data.append({
            'Matrícula': maestro.id_usuario.matricula,
            'Nombre': maestro.id_usuario.nombre,
            'Apellido': maestro.id_usuario.apellido,
            'Rol': 'Maestro',
            'Departamento': maestro.departamento,
            'Cubículo': maestro.cubiculo or '',
            'Grado Académico': maestro.grado_academico,
        })

    for administrativo in Administrativos.objects.select_related('id_usuario').all():
        usuarios_data.append({
            'Matrícula': administrativo.id_usuario.matricula,
            'Nombre': administrativo.id_usuario.nombre,
            'Apellido': administrativo.id_usuario.apellido,
            'Rol': 'Administrativo',
            'Departamento': administrativo.departamento,
            'Puesto': administrativo.puesto,
        })

    for administrador in Administrador.objects.select_related('id_usuario').all():
        usuarios_data.append({
            'Matrícula': administrador.id_usuario.matricula,
            'Nombre': administrador.id_usuario.nombre,
            'Apellido': administrador.id_usuario.apellido,
            'Rol': 'Administrador',
            'Puesto': administrador.puesto,
            'Nivel Prioridad': administrador.nivel_prioridad,
        })

    usuarios_data.sort(key=lambda x: x['Matrícula'])
    return usuarios_data


def _filtrar_usuarios_exportacion(
    usuarios_data: list[dict],
    *,
    busqueda: str,
    rol_filtro: str,
) -> list[dict]:
    if busqueda:
        texto_busqueda = _normalizar_texto(busqueda)

        def coincide(usuario: dict) -> bool:
            campos = [
                usuario.get('Matrícula', ''),
                usuario.get('Nombre', ''),
                usuario.get('Apellido', ''),
                f"{usuario.get('Nombre', '')} {usuario.get('Apellido', '')}",
                usuario.get('Rol', ''),
                usuario.get('Carrera', ''),
                usuario.get('Departamento', ''),
                usuario.get('Puesto', ''),
                usuario.get('Grado Académico', ''),
                str(usuario.get('Semestre', '')),
                str(usuario.get('Nivel Prioridad', '')),
                usuario.get('Estatus', ''),
            ]
            return any(texto_busqueda in _normalizar_texto(campo) for campo in campos if campo is not None)

        usuarios_data = [usuario for usuario in usuarios_data if coincide(usuario)]

    if rol_filtro:
        rol_etiqueta = _ROL_MAP_EXPORTACION.get(rol_filtro, rol_filtro)
        usuarios_data = [usuario for usuario in usuarios_data if usuario.get('Rol') == rol_etiqueta]

    return usuarios_data


def _excel_alineacion_celda(valor, columna: str) -> str:
    if _pdf_es_celda_vacia(str(valor)):
        return 'center'
    return _ALINEACION_EXCEL_USUARIOS.get(columna, 'center')


_EXCEL_MARGEN_FILAS_SUPERIOR = 2
_EXCEL_MARGEN_COL_IZQUIERDA = 1
_EXCEL_FILA_CONTENIDO = _EXCEL_MARGEN_FILAS_SUPERIOR + 1
_EXCEL_COL_CONTENIDO = _EXCEL_MARGEN_COL_IZQUIERDA + 1


def _excel_celda_datos(col_idx: int) -> int:
    return _EXCEL_COL_CONTENIDO + col_idx - 1


def _excel_aplicar_margenes_hoja(worksheet) -> None:
    from openpyxl.utils import get_column_letter

    worksheet.column_dimensions[get_column_letter(_EXCEL_MARGEN_COL_IZQUIERDA)].width = 4
    for fila in range(1, _EXCEL_MARGEN_FILAS_SUPERIOR + 1):
        worksheet.row_dimensions[fila].height = 20


def _excel_insertar_logo(worksheet, *, ancla: str) -> bool:
    logo_ruta = _pdf_ruta_logo()
    if not logo_ruta:
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage

        imagen = XLImage(logo_ruta)
        imagen.width = 72
        imagen.height = 72
        worksheet.add_image(imagen, ancla)
        return True
    except Exception:
        logger.warning('No se pudo cargar el logo para el Excel de usuarios', exc_info=True)
        return False


def _excel_dibujar_cabecera_documento(
    worksheet,
    *,
    num_columnas: int,
    fecha: datetime,
    total_registros: int,
    filtros: list[str] | None = None,
) -> int:
    """Replica la cabecera del PDF con márgenes, logo, textos y línea separadora."""
    from openpyxl.utils import get_column_letter

    _excel_aplicar_margenes_hoja(worksheet)

    col_logo = get_column_letter(_EXCEL_COL_CONTENIDO)
    ancla_logo = f'{col_logo}{_EXCEL_FILA_CONTENIDO}'
    tiene_logo = _excel_insertar_logo(worksheet, ancla=ancla_logo)
    col_texto_idx = _EXCEL_COL_CONTENIDO + (1 if tiene_logo else 0)
    col_texto = get_column_letter(col_texto_idx)
    end_col_letter = get_column_letter(_excel_celda_datos(num_columnas))
    fila = _EXCEL_FILA_CONTENIDO

    title_font = Font(name='Arial', size=18, bold=True, color='111827')
    subtitle_font = Font(name='Arial', size=10, color='6B7280')
    meta_font = Font(name='Arial', size=9, color='6B7280')
    filtros_font = Font(name='Arial', size=8, italic=True, color='6B7280')
    separador_border = Border(
        bottom=Side(style='thin', color='D1D5DB'),
    )

    if tiene_logo:
        worksheet.column_dimensions[col_logo].width = 12
        worksheet.row_dimensions[fila].height = 22
        worksheet.row_dimensions[fila + 1].height = 16
        worksheet.row_dimensions[fila + 2].height = 16
        worksheet.row_dimensions[fila + 3].height = 14

    worksheet[f'{col_texto}{fila}'] = 'Reporte de Usuarios'
    worksheet[f'{col_texto}{fila}'].font = title_font
    worksheet.merge_cells(f'{col_texto}{fila}:{end_col_letter}{fila}')
    worksheet[f'{col_texto}{fila}'].alignment = Alignment(horizontal='left', vertical='center')

    worksheet[f'{col_texto}{fila + 1}'] = 'SchoolTrack · Sistema de gestión escolar'
    worksheet[f'{col_texto}{fila + 1}'].font = subtitle_font
    worksheet.merge_cells(f'{col_texto}{fila + 1}:{end_col_letter}{fila + 1}')
    worksheet[f'{col_texto}{fila + 1}'].alignment = Alignment(horizontal='left', vertical='center')

    worksheet[f'{col_texto}{fila + 2}'] = (
        f'Exportado el {_pdf_fecha_legible(fecha)} · {total_registros} registro(s)'
    )
    worksheet[f'{col_texto}{fila + 2}'].font = meta_font
    worksheet.merge_cells(f'{col_texto}{fila + 2}:{end_col_letter}{fila + 2}')
    worksheet[f'{col_texto}{fila + 2}'].alignment = Alignment(horizontal='left', vertical='center')

    fila_separador = fila + 3
    if filtros:
        worksheet[f'{col_texto}{fila_separador}'] = 'Filtros: ' + ' · '.join(filtros)
        worksheet[f'{col_texto}{fila_separador}'].font = filtros_font
        worksheet.merge_cells(f'{col_texto}{fila_separador}:{end_col_letter}{fila_separador}')
        worksheet[f'{col_texto}{fila_separador}'].alignment = Alignment(horizontal='left', vertical='center')
        fila_separador += 1

    for col_num in range(_EXCEL_COL_CONTENIDO, _excel_celda_datos(num_columnas) + 1):
        cell = worksheet.cell(row=fila_separador, column=col_num)
        cell.border = separador_border

    return fila_separador + 2


def _excel_generar_workbook_usuarios(
    usuarios_data: list[dict],
    *,
    rol_filtro: str,
    busqueda: str,
    ahora: datetime,
) -> BytesIO:
    from openpyxl.utils import get_column_letter

    headers = _pdf_headers_usuarios(rol_filtro)
    filas = [_pdf_fila_usuario_exportacion(usuario, rol_filtro) for usuario in usuarios_data]

    filtros_activos = []
    if busqueda:
        filtros_activos.append(f'Búsqueda: {busqueda}')
    if rol_filtro:
        filtros_activos.append(f'Rol: {_ROL_MAP_EXPORTACION.get(rol_filtro, rol_filtro)}')

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        worksheet = writer.book.create_sheet('Usuarios')
        writer.sheets['Usuarios'] = worksheet
        if 'Sheet1' in writer.book.sheetnames:
            del writer.book['Sheet1']

        fila_encabezados = _excel_dibujar_cabecera_documento(
            worksheet,
            num_columnas=len(headers),
            fecha=ahora,
            total_registros=len(usuarios_data),
            filtros=filtros_activos or None,
        )

        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=9, color='1F2937')
        na_font = Font(name='Arial', size=9, color='BEC4D0')
        footer_font = Font(name='Arial', size=8, italic=True, color='9CA3AF')

        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        border_color = 'D1D5DB'
        thin_border = Border(
            left=Side(style='thin', color=border_color),
            right=Side(style='thin', color=border_color),
            top=Side(style='thin', color=border_color),
            bottom=Side(style='thin', color=border_color),
        )

        for col_num, column in enumerate(headers, 1):
            col_excel = _excel_celda_datos(col_num)
            cell = worksheet.cell(row=fila_encabezados, column=col_excel, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for row_offset, row in enumerate(filas):
            fila_excel = fila_encabezados + 1 + row_offset
            es_zebra = row_offset % 2 == 1
            for col_num, value in enumerate(row, 1):
                col_excel = _excel_celda_datos(col_num)
                cell = worksheet.cell(row=fila_excel, column=col_excel)
                columna = headers[col_num - 1]
                texto = '' if pd.isna(value) else str(value)
                cell.value = texto
                cell.font = na_font if texto == _PDF_VALOR_NO_APLICA else data_font
                cell.alignment = Alignment(
                    horizontal=_excel_alineacion_celda(texto, columna),
                    vertical='center',
                    wrap_text=columna in ('Nombre Completo', 'Carrera', 'Departamento'),
                )
                cell.border = thin_border
                if es_zebra:
                    cell.fill = zebra_fill

        for col_num, column in enumerate(headers, 1):
            minimo, maximo = _ANCHO_EXCEL_USUARIOS.get(column, (12, 30))
            largo_maximo = len(column)
            for fila in filas:
                if col_num - 1 < len(fila):
                    largo_maximo = max(largo_maximo, len(str(fila[col_num - 1])))
            worksheet.column_dimensions[get_column_letter(_excel_celda_datos(col_num))].width = max(
                minimo,
                min(maximo, largo_maximo + 2),
            )

        col_pie = get_column_letter(_EXCEL_COL_CONTENIDO)
        ultima_fila = fila_encabezados + len(filas) + 2
        worksheet[f'{col_pie}{ultima_fila}'] = 'Generado automáticamente por SchoolTrack'
        worksheet[f'{col_pie}{ultima_fila}'].font = footer_font
        worksheet[f'{col_pie}{ultima_fila}'].alignment = Alignment(horizontal='left', vertical='center')

        from openpyxl.worksheet.header_footer import HeaderFooterItem, _HeaderFooterPart

        worksheet.print_title_rows = f'{fila_encabezados}:{fila_encabezados}'
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.6
        worksheet.page_margins.right = 0.6
        worksheet.page_margins.top = 0.6
        worksheet.page_margins.bottom = 0.6
        pie_pagina = HeaderFooterItem()
        pie_pagina.left = _HeaderFooterPart('Generado automáticamente por SchoolTrack')
        pie_pagina.right = _HeaderFooterPart('Página &P')
        worksheet.oddFooter = pie_pagina

    output.seek(0)
    return output


def sesion_roles_permitidas(request, roles: tuple) -> bool:
    """Usado en todo el CRUD: verifica request.session['usuario_rol'] contra roles permitidos."""
    role = request.session.get('usuario_rol')
    return role is not None and role in roles


def construir_direccion(data) -> str | None:
    """
    Une los campos separados de dirección en un solo texto para guardar en DatosPersonales.direccion.
    """
    direccion_directa = data.get('direccion', '').strip()
    if direccion_directa:
        return direccion_directa

    partes = [
        data.get('calle', '').strip(),
        data.get('numero_exterior', '').strip(),
        data.get('numero_interior', '').strip(),
        data.get('colonia', '').strip(),
        data.get('municipio', '').strip(),
        data.get('estado', '').strip(),
        data.get('cp', '').strip(),
    ]

    partes = [parte for parte in partes if parte]
    if not partes:
        return None

    return ', '.join(partes)


def _parse_fecha_nacimiento(valor: str):
    if not valor:
        return None
    return datetime.strptime(valor, '%Y-%m-%d').date()


def _periodo_actual() -> str:
    hoy = timezone.now()
    if hoy.month <= 6:
        periodo = 'A'
    elif hoy.month >= 8:
        periodo = 'B'
    else:
        # Julio queda como A para no romper el flujo de captura.
        periodo = 'A'

    return f"{hoy.year}-{periodo}"


def desglosar_direccion(direccion: str | None) -> dict:
    """Convierte la dirección guardada en texto a campos separados para editarla."""
    if not direccion:
        return {
            'calle': '',
            'numero_exterior': '',
            'numero_interior': '',
            'colonia': '',
            'municipio': '',
            'estado': '',
            'cp': '',
        }

    partes = [parte.strip() for parte in str(direccion).split(',')]
    partes += [''] * (7 - len(partes))

    return {
        'calle': partes[0],
        'numero_exterior': partes[1],
        'numero_interior': partes[2],
        'colonia': partes[3],
        'municipio': partes[4],
        'estado': partes[5],
        'cp': partes[6],
    }


_FILTROS_LISTA_USUARIOS = ('rol', 'q', 'page')


def _filtros_lista_usuarios(request) -> dict:
    """Lee rol, búsqueda y página desde GET o POST (campos ocultos al editar)."""
    origen = request.POST if request.method == 'POST' else request.GET
    filtros = {clave: origen.get(clave, '').strip() for clave in _FILTROS_LISTA_USUARIOS}
    return {clave: valor for clave, valor in filtros.items() if valor}


def _query_lista_usuarios(filtros: dict) -> str:
    return '?' + urlencode(filtros) if filtros else ''


def _anexar_retorno_lista_usuarios(request, context: dict) -> dict:
    filtros = _filtros_lista_usuarios(request)
    context['retorno_filtros'] = filtros
    context['url_lista_usuarios'] = reverse('gestion_usuarios') + _query_lista_usuarios(filtros)
    return context


def _redirect_gestion_usuarios(request):
    return redirect(reverse('gestion_usuarios') + _query_lista_usuarios(_filtros_lista_usuarios(request)))


def _validar_foto_perfil(foto_archivo) -> str | None:
    """Valida dimensiones y tamaño de la foto de perfil. Devuelve mensaje de error o None."""
    from PIL import Image

    img = Image.open(foto_archivo)
    width, height = img.size
    if width != 480 or height != 640:
        return f'La foto debe tener dimensiones de 480x640 píxeles. La foto subida tiene {width}x{height} píxeles.'
    if foto_archivo.size > 5 * 1024 * 1024:
        return 'La foto no puede superar los 5MB.'
    return None


def _quitar_foto_perfil(usuario: Usuarios) -> None:
    """Elimina el archivo de foto y deja el campo vacío en el usuario."""
    if usuario.foto:
        usuario.foto.delete(save=False)
        usuario.foto = None


def _normalizar_texto(valor: str) -> str:
    texto = unicodedata.normalize('NFD', valor or '')
    texto = ''.join(ch for ch in texto if unicodedata.category(ch) != 'Mn')
    return texto.lower().strip()


def _validar_datos_personales_unicos(correo: str, curp: str, usuario_id: int | None = None) -> str | None:
    """
    Valida solo los datos que deben ser únicos en el sistema.
    Devuelve un mensaje de error si encuentra duplicados; si no, devuelve None.
    """
    errores = []

    if correo:
        correos = DatosPersonales.objects.filter(correo_inst__iexact=correo)
        if usuario_id is not None:
            correos = correos.exclude(id_usuario_id=usuario_id)
        if correos.exists():
            errores.append('El correo ya está registrado')

    if curp:
        curps = DatosPersonales.objects.filter(curp__iexact=curp)
        if usuario_id is not None:
            curps = curps.exclude(id_usuario_id=usuario_id)
        if curps.exists():
            errores.append('La CURP ya está registrada')

    if errores:
        return '. '.join(errores) + '.'

    return None


def _formatear_error_validacion(error: ValidationError) -> str:
    """
    Convierte un ValidationError del modelo en un mensaje de error legible.
    """
    if hasattr(error, 'message_dict'):
        # Error con múltiples campos
        mensajes = []
        for campo, errores in error.message_dict.items():
            if isinstance(errores, list):
                mensajes.extend(errores)
            else:
                mensajes.append(str(errores))
        return '. '.join(mensajes)
    elif hasattr(error, 'messages'):
        # Error con lista de mensajes
        return '. '.join(error.messages)
    else:
        # Error simple
        return str(error)


def _validar_nombre_carrera_unico(nombre: str, carrera_id: int | None = None) -> str | None:
    """
    Evita duplicados de carrera por nombre, ignorando mayúsculas, minúsculas y acentos.
    """
    nombre_normalizado = _normalizar_texto(nombre)
    if not nombre_normalizado:
        return None

    for carrera in Carrera.objects.all(): #valida carreras duplicadas
        if carrera_id is not None and carrera.id == carrera_id:
            continue

        if _normalizar_texto(carrera.nombre) == nombre_normalizado:
            return 'Ya existe una carrera con ese nombre.'

    return None


# ==================== VISTAS PRINCIPALES DE ADMINISTRADOR ====================
# CRUD gestión de usuarios (login/urls.py → admin_views.py → Templates/administrador/):
#   R → gestion_usuarios      + GestionUsuarios.html
#   C → agregar_usuario       + AgregarUsuario.html
#   U → editar_usuario        + EditarUsuario.html
#   D → eliminar_usuario      + modal AJAX en GestionUsuarios.html
# Extra: restablecer_contrasena, exportar_usuarios, exportar_usuarios_pdf

def gestion_usuarios(request):
    """
    READ (R del CRUD): listar y consultar usuarios.
    URL: /administrador/usuarios/  (name='gestion_usuarios')
    Template: administrador/GestionUsuarios.html

    Flujo:
      1. Consulta las 4 tablas por rol y arma una lista unificada (usuarios_data)
      2. Filtra por búsqueda (?q=) y por rol (?rol=)
      3. Pagina de 10 en 10 (?page=)
      4. Renderiza la tabla con botones Editar / Eliminar / Restablecer
    """
    # --- PASO 1: SEGURIDAD — solo administradores ---
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')
    
    # --- PASO 2: CONSULTAR BD — un dict por usuario para la tabla HTML ---
    usuarios_data = []
    
    # select_related: trae id_usuario e id_carrera en la misma query (más eficiente)
    alumnos = Alumnos.objects.select_related('id_usuario', 'id_carrera').all()
    for alumno in alumnos:
        usuarios_data.append({
            'id_usuario': alumno.id_usuario.id_usuario,
            'matricula': alumno.id_usuario.matricula,
            'nombre': alumno.id_usuario.nombre,
            'apellido': alumno.id_usuario.apellido,
            'rol': 'alumno',
            'foto_url': alumno.id_usuario.foto.url if alumno.id_usuario.foto else '',
            'carrera': str(alumno.id_carrera) if alumno.id_carrera else '',
            'semestre': alumno.semestre,
            'estatus': alumno.estatus,
            'ultimo_acceso': alumno.id_usuario.ultimo_acceso
        })
    
    # Maestros — mismo patrón: dict unificado para la tabla Read
    maestros = Maestros.objects.select_related('id_usuario').all()
    for maestro in maestros:
        usuarios_data.append({
            'id_usuario': maestro.id_usuario.id_usuario,
            'matricula': maestro.id_usuario.matricula,
            'nombre': maestro.id_usuario.nombre,
            'apellido': maestro.id_usuario.apellido,
            'rol': 'maestro',
            'foto_url': maestro.id_usuario.foto.url if maestro.id_usuario.foto else '',
            'departamento': maestro.departamento,
            'cubiculo': maestro.cubiculo,
            'grado_academico': maestro.grado_academico,
            'ultimo_acceso': maestro.id_usuario.ultimo_acceso
        })
    
    # Administrativos
    administrativos = Administrativos.objects.select_related('id_usuario').all()
    for administrativo in administrativos:
        usuarios_data.append({
            'id_usuario': administrativo.id_usuario.id_usuario,
            'matricula': administrativo.id_usuario.matricula,
            'nombre': administrativo.id_usuario.nombre,
            'apellido': administrativo.id_usuario.apellido,
            'rol': 'administrativo',
            'foto_url': administrativo.id_usuario.foto.url if administrativo.id_usuario.foto else '',
            'departamento': administrativo.departamento,
            'puesto': administrativo.puesto,
            'ultimo_acceso': administrativo.id_usuario.ultimo_acceso
        })
    
    # Administradores
    administradores = Administrador.objects.select_related('id_usuario').all()
    for administrador in administradores:
        usuarios_data.append({
            'id_usuario': administrador.id_usuario.id_usuario,
            'matricula': administrador.id_usuario.matricula,
            'nombre': administrador.id_usuario.nombre,
            'apellido': administrador.id_usuario.apellido,
            'rol': 'admin',
            'foto_url': administrador.id_usuario.foto.url if administrador.id_usuario.foto else '',
            'puesto': administrador.puesto,
            'nivel_prioridad': administrador.nivel_prioridad,
            'ultimo_acceso': administrador.id_usuario.ultimo_acceso
        })
    
    # --- PASO 3: ORDENAR la lista completa por id_usuario ---
    usuarios_data.sort(key=lambda x: x['id_usuario'])

    # --- PASO 4: FILTRO DE BÚSQUEDA (?q=texto) — request.GET, no POST ---
    busqueda = request.GET.get('q', '').strip()
    usuarios_filtrados = usuarios_data
    if busqueda:
        texto_busqueda = _normalizar_texto(busqueda)

        def coincide(usuario: dict) -> bool:
            campos = [
                str(usuario.get('id_usuario', '')),
                usuario.get('matricula', ''),
                usuario.get('nombre', ''),
                usuario.get('apellido', ''),
                f"{usuario.get('nombre', '')} {usuario.get('apellido', '')}",  # Nombre completo
                usuario.get('rol', ''),
                usuario.get('carrera', ''),
                usuario.get('departamento', ''),
                usuario.get('puesto', ''),
                usuario.get('grado_academico', ''),
                str(usuario.get('semestre', '')),
                str(usuario.get('nivel_prioridad', '')),
                usuario.get('estatus', ''),
            ]
            return any(texto_busqueda in _normalizar_texto(campo) for campo in campos if campo is not None)

        usuarios_filtrados = [usuario for usuario in usuarios_data if coincide(usuario)]

    # --- PASO 5: FILTRO POR ROL (?rol=alumno|maestro|...) ---
    rol_filtro = request.GET.get('rol', '').strip()
    if rol_filtro:
        usuarios_filtrados = [usuario for usuario in usuarios_filtrados if usuario.get('rol') == rol_filtro]
    
    # --- PASO 6: PAGINACIÓN — 10 usuarios por página (?page=2) ---
    paginator = Paginator(usuarios_filtrados, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # --- PASO 7: ENVIAR DATOS AL TEMPLATE — {% for usuario in usuarios %} ---
    filtros_actuales = _filtros_lista_usuarios(request)
    if page_obj.number > 1:
        filtros_actuales.setdefault('page', str(page_obj.number))

    context = {
        'usuarios': page_obj,
        'total_usuarios': len(usuarios_data),
        'usuarios_encontrados': len(usuarios_filtrados),
        'busqueda': busqueda,
        'rol_filtro': rol_filtro,
        'retorno_filtros': filtros_actuales,
        'query_lista': _query_lista_usuarios(filtros_actuales),
        'perfil': {
            'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
            'matricula': request.session.get('usuario_matricula', 'N/A')
        }
    }
    
    return render(request, 'administrador/GestionUsuarios.html', context)


def exportar_usuarios(request):
    """Exporta la lista de usuarios a Excel con el mismo formato que el reporte PDF."""
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')

    busqueda = request.GET.get('q', '').strip()
    rol_filtro = request.GET.get('rol', '').strip()
    usuarios_data = _filtrar_usuarios_exportacion(
        _recolectar_usuarios_exportacion(),
        busqueda=busqueda,
        rol_filtro=rol_filtro,
    )

    output = _excel_generar_workbook_usuarios(
        usuarios_data,
        rol_filtro=rol_filtro,
        busqueda=busqueda,
        ahora=datetime.now(),
    )

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios_{}.xlsx"'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )
    return response


def exportar_usuarios_pdf(request):
    """Exporta la lista de usuarios a PDF con formato profesional"""
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')

    busqueda = request.GET.get('q', '').strip()
    rol_filtro = request.GET.get('rol', '').strip()
    usuarios_data = _filtrar_usuarios_exportacion(
        _recolectar_usuarios_exportacion(),
        busqueda=busqueda,
        rol_filtro=rol_filtro,
    )

    ahora = datetime.now()
    filtros_activos = []
    if busqueda:
        filtros_activos.append(f'Búsqueda: {busqueda}')
    if rol_filtro:
        filtros_activos.append(f'Rol: {_ROL_MAP_EXPORTACION.get(rol_filtro, rol_filtro)}')

    pdf = _ReporteUsuariosPDF(orientation='L', unit='mm', format='A4')
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    _pdf_dibujar_cabecera_documento(
        pdf,
        fecha=ahora,
        total_registros=len(usuarios_data),
        filtros=filtros_activos or None,
    )

    headers = _pdf_headers_usuarios(rol_filtro)
    filas = [_pdf_fila_usuario_exportacion(usuario, rol_filtro) for usuario in usuarios_data]
    x_tabla, col_widths = _pdf_layout_tabla_fluida(pdf, headers, filas)
    alineaciones = _pdf_alineaciones_columnas(headers)
    _pdf_dibujar_encabezados_tabla(pdf, headers, col_widths, x_inicio=x_tabla)

    for indice, fila in enumerate(filas):
        _pdf_dibujar_fila_tabla(
            pdf,
            fila,
            col_widths,
            alineaciones=alineaciones,
            x_inicio=x_tabla,
            fill=indice % 2 == 1,
            headers=headers,
        )

    # Generar respuesta
    response = HttpResponse(_fpdf_output_bytes(pdf), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios_{}.pdf"'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )

    return response


def gestion_carreras(request):
    """Vista para administrar carreras"""
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')

    context = {
        'carreras': Carrera.objects.order_by('nombre').all(),
        'perfil': {
            'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
            'matricula': request.session.get('usuario_matricula', 'N/A')
        }
    }
    return render(request, 'administrador/GestionCarreras.html', context)


def agregar_carrera(request):
    """Crea una nueva carrera"""
    if not sesion_roles_permitidas(request, ('admin',)):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        return redirect('selector_rol')

    if request.method != 'POST':
        return redirect('gestion_carreras')

    es_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    try:
        nombre = request.POST.get('nombre', '').strip()
        clave = request.POST.get('clave', '').strip().upper()

        if not nombre or not clave:
            error_msg = 'Nombre y clave son obligatorios'
            if es_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('gestion_carreras')

        error_nombre = _validar_nombre_carrera_unico(nombre)
        if error_nombre:
            if es_ajax:
                return JsonResponse({'success': False, 'error': error_nombre})
            messages.error(request, error_nombre)
            return redirect('gestion_carreras')

        if Carrera.objects.filter(clave__iexact=clave).exists():
            error_msg = 'Ya existe una carrera con esa clave'
            if es_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('gestion_carreras')

        Carrera.objects.create(nombre=nombre, clave=clave)
        success_msg = f'Carrera {nombre} agregada correctamente'
        if es_ajax:
            return JsonResponse({'success': True, 'message': success_msg})
        messages.success(request, success_msg)
    except Exception as e:
        error_msg = f'Error al agregar carrera: {str(e)}'
        if es_ajax:
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)

    return redirect('gestion_carreras')


def editar_carrera(request, carrera_id):
    """Edita una carrera existente"""
    if not sesion_roles_permitidas(request, ('admin',)):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        return redirect('selector_rol')

    carrera = get_object_or_404(Carrera, pk=carrera_id)

    if request.method != 'POST':
        return redirect('gestion_carreras')

    try:
        nombre = request.POST.get('nombre', '').strip()
        clave = request.POST.get('clave', '').strip().upper()

        if not nombre or not clave:
            error_msg = 'Nombre y clave son obligatorios'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('gestion_carreras')

        error_nombre = _validar_nombre_carrera_unico(nombre, carrera.id)
        if error_nombre:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_nombre})
            messages.error(request, error_nombre)
            return redirect('gestion_carreras')

        if Carrera.objects.exclude(pk=carrera.pk).filter(clave__iexact=clave).exists():
            error_msg = 'Ya existe otra carrera con esa clave'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('gestion_carreras')

        # Advertencia si la carrera tiene alumnos
        num_alumnos = Alumnos.objects.filter(id_carrera=carrera).count()
        warning_msg = None
        if num_alumnos > 0:
            warning_msg = f'Esta carrera tiene {num_alumnos} alumno(s) inscrito(s). Cambiar el nombre puede afectar reportes históricos.'

        carrera.nombre = nombre
        carrera.clave = clave
        carrera.save()

        success_msg = f'Carrera {nombre} actualizada correctamente'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': success_msg, 'warning': warning_msg})
        messages.success(request, success_msg)
        if warning_msg:
            messages.warning(request, warning_msg)
    except Exception as e:
        error_msg = f'Error al editar carrera: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)

    return redirect('gestion_carreras')


def eliminar_carrera(request, carrera_id):
    """Elimina una carrera"""
    if not sesion_roles_permitidas(request, ('admin',)):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        return redirect('selector_rol')

    if request.method != 'POST':
        return redirect('gestion_carreras')

    carrera = get_object_or_404(Carrera, pk=carrera_id)

    try:
        # Validar que la carrera no tenga alumnos asociados
        if Alumnos.objects.filter(id_carrera=carrera).exists():
            error_msg = f'No se puede eliminar "{carrera.nombre}" porque tiene alumnos inscritos'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('gestion_carreras')

        # Validar que la carrera no tenga grupos asociados (con materias)
        if Grupo.objects.filter(id_carrera=carrera).exists():
            error_msg = f'No se puede eliminar "{carrera.nombre}" porque tiene grupos/materias asignadas'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('gestion_carreras')

        nombre = carrera.nombre
        carrera.delete()
        success_msg = f'Carrera {nombre} eliminada correctamente'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': success_msg})
        messages.success(request, success_msg)
    except Exception as e:
        error_msg = f'Error al eliminar carrera: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)

    return redirect('gestion_carreras')


def verificar_clave_carrera(request):
    """Verifica si una clave de carrera ya existe (para validación AJAX)"""
    if not sesion_roles_permitidas(request, ('admin',)):
        return JsonResponse({'existe': False}, status=403)

    clave = request.GET.get('clave', '').strip().upper()
    carrera_id = request.GET.get('carrera_id', '')

    if not clave:
        return JsonResponse({'existe': False})

    queryset = Carrera.objects.filter(clave__iexact=clave)

    # Si estamos editando, excluir la carrera actual
    if carrera_id:
        try:
            queryset = queryset.exclude(pk=int(carrera_id))
        except ValueError:
            pass

    existe = queryset.exists()
    return JsonResponse({'existe': existe})


def verificar_nombre_carrera(request):
    """Verifica si un nombre de carrera ya existe (para validación AJAX)"""
    if not sesion_roles_permitidas(request, ('admin',)):
        return JsonResponse({'existe': False}, status=403)

    nombre = request.GET.get('nombre', '').strip()
    carrera_id = request.GET.get('carrera_id', '')

    if not nombre:
        return JsonResponse({'existe': False})

    exclude_id = None
    if carrera_id:
        try:
            exclude_id = int(carrera_id)
        except ValueError:
            pass

    error = _validar_nombre_carrera_unico(nombre, exclude_id)
    return JsonResponse({
        'existe': bool(error),
        'mensaje': error or '',
    })


def agregar_usuario(request):
    """
    CREATE (C del CRUD): alta de un usuario nuevo.
    URL: /administrador/usuarios/agregar/  (name='agregar_usuario' en urls.py)
    Template: administrador/AgregarUsuario.html

    Flujo general:
      GET  → muestra formulario vacío (líneas ~1206-1220)
      POST → lee formulario → valida → guarda en BD → redirect a gestion_usuarios
    """
    # --- PASO 1: SEGURIDAD — solo administradores pueden crear usuarios ---
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')
    
    # --- PASO 2: DATOS INFORMATIVOS PARA EL FORMULARIO (no crean nada aún) ---
    # proximo_id: vista previa del ID que asignará la secuencia de PostgreSQL al guardar
    proximo_id = _proximo_id_usuario_preview()
    
    # siguientes: preview de matrícula por rol (ej. EMP-20260004). La matrícula REAL
    # se genera en Usuarios.save() del modelo (models.py), no aquí.
    año_actual = timezone.now().year
    siguientes = {}
    prefijos = {'admin': 'ADM-', 'maestro': 'EMP-', 'administrativo': 'AD-', 'alumno': ''}
    
    for rol, prefijo in prefijos.items():
        patron = f"{prefijo}{año_actual}"
        ultima = Usuarios.objects.filter(matricula__startswith=patron).order_by('-matricula').first()
        if ultima:
            try:
                ultimo_num = int(ultima.matricula[-4:])
                siguientes[rol] = f"{patron}{ultimo_num + 1:04d}"
            except:
                siguientes[rol] = f"{patron}0001"
        else:
            siguientes[rol] = f"{patron}0001"
    
    # --- PASO 3: ¿EL ADMIN ENVIÓ EL FORMULARIO? (POST = Create) ---
    if request.method == 'POST':
        # --- PASO 3a: LEER CADA CAMPO DEL HTML (name="..." del formulario) ---
        # request.POST.get('nombre') ← coincide con <input name="nombre"> en AgregarUsuario.html
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        rol = request.POST.get('rol', '').strip()
        if rol == 'administrador':
            rol = 'admin'  # en BD el valor guardado es 'admin', no 'administrador'
        
        # Contraseña aleatoria segura (password_utils.py). Se encripta al hacer usuario.save()
        contrasena_temporal = generar_contrasena_temporal()
        
        # Campos opcionales / por rol (algunos los inserta JS dinámico según el select de rol)
        correo = request.POST.get('correo_inst', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        curp = request.POST.get('curp', '').strip()
        fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()
        genero = request.POST.get('genero', '').strip()
        carrera_id = request.POST.get('carrera_id', '').strip()
        semestre = request.POST.get('semestre', '').strip()
        departamento = request.POST.get('departamento', '').strip()
        cubiculo = request.POST.get('cubiculo', '').strip()
        grado_academico = request.POST.get('grado_academico', '').strip()
        puesto = request.POST.get('puesto', '').strip()
        nivel_prioridad = request.POST.get('nivel_prioridad', '').strip()
        id_ciclo_escolar = request.POST.get('id_ciclo_escolar', '').strip()
        estatus = request.POST.get('estatus', 'Activo').strip()
        # Une calle, colonia, cp, etc. en un solo texto para DatosPersonales.direccion
        direccion = construir_direccion(request.POST)
        
        # --- PASO 3b: VALIDACIONES (si falla algo, NO se toca la BD) ---
        # Clave = nombre del campo HTML; valor = mensaje de error para mostrar en rojo
        errores_campos = {}
        
        # ============ VALIDAR CAMPOS OBLIGATORIOS ============
        if not nombre:
            errores_campos['nombre'] = 'El nombre es obligatorio'
        elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$', nombre):
            errores_campos['nombre'] = 'El nombre solo puede contener letras y espacios'
        
        if not apellido:
            errores_campos['apellido'] = 'El apellido es obligatorio'
        elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$', apellido):
            errores_campos['apellido'] = 'El apellido solo puede contener letras y espacios'
        
        if not rol:
            errores_campos['rol'] = 'El rol es obligatorio'
        
        # ============ VALIDAR CAMPOS ESPECÍFICOS POR ROL ============
        if rol == 'alumno' and not carrera_id:
            errores_campos['carrera_id'] = 'Por favor selecciona una carrera de la lista para el alumno'
        
        if rol == 'alumno' and carrera_id:
            try:
                Carrera.objects.get(pk=carrera_id)
            except Carrera.DoesNotExist:
                errores_campos['carrera_id'] = 'La carrera seleccionada no es válida. Por favor selecciona otra'
        
        if rol == 'alumno' and not semestre:
            errores_campos['semestre'] = 'Por favor ingresa el semestre actual del alumno (1-12)'
        
        if rol == 'alumno' and semestre:
            try:
                semestre_int = int(semestre)
                if semestre_int < 1 or semestre_int > 12:
                    errores_campos['semestre'] = 'El semestre debe ser un número entre 1 y 12'
            except ValueError:
                errores_campos['semestre'] = 'El semestre debe ser un número válido (ejemplo: 3)'
        
        if rol == 'alumno' and estatus:
            estatus_validos = ['Activo', 'Baja', 'Egresado']
            if estatus not in estatus_validos:
                errores_campos['estatus'] = 'El estatus seleccionado no es válido'
        
        if rol == 'maestro' and not grado_academico:
            errores_campos['grado_academico'] = 'Por favor selecciona el grado académico del maestro'
        
        if rol == 'maestro' and grado_academico:
            grados_validos = ['Licenciatura', 'Maestria', 'Doctorado']
            if grado_academico not in grados_validos:
                errores_campos['grado_academico'] = 'Por favor selecciona un grado académico válido'
        
        if rol == 'maestro' and not departamento:
            errores_campos['departamento'] = 'Por favor ingresa el departamento donde trabaja el maestro'
        
        if rol == 'administrativo' and not puesto:
            errores_campos['puesto'] = 'Por favor ingresa el puesto del administrativo'
        
        if rol == 'administrativo' and not departamento:
            errores_campos['departamento'] = 'Por favor ingresa el departamento del administrativo'
        
        if rol == 'admin' and not puesto:
            errores_campos['puesto'] = 'Por favor selecciona el puesto del administrador'
        
        if rol == 'admin' and puesto:
            puestos_validos = ['Director', 'Subdirector', 'Auxiliar']
            if puesto not in puestos_validos:
                errores_campos['puesto'] = 'Por favor selecciona un puesto válido de la lista'
        
        if rol == 'admin' and nivel_prioridad:
            try:
                prioridad_int = int(nivel_prioridad)
                if prioridad_int < 1 or prioridad_int > 10:
                    errores_campos['nivel_prioridad'] = 'El nivel de prioridad debe ser un número entre 1 y 10'
            except ValueError:
                errores_campos['nivel_prioridad'] = 'El nivel de prioridad debe ser un número válido (ejemplo: 5)'
        
        # ============ VALIDAR DATOS PERSONALES ============
        # Validar teléfono (solo números y 10 dígitos si se proporciona)
        if telefono:
            telefono_limpio = telefono.replace('-', '').replace(' ', '').replace('(', '').replace(')', '')
            if not telefono_limpio.isdigit():
                errores_campos['telefono'] = 'El teléfono debe contener solo números (sin guiones ni espacios)'
            elif len(telefono_limpio) != 10:
                errores_campos['telefono'] = 'El teléfono debe tener exactamente 10 dígitos'
        
        # Validar correo duplicado y formato
        if correo:
            # Validar formato de email usando EmailValidator de Django
            try:
                validate_email(correo)
            except ValidationError:
                errores_campos['correo_inst'] = 'El formato del correo no es válido.'
            else:
                # Si el formato es válido, verificar duplicados
                if DatosPersonales.objects.filter(correo_inst__iexact=correo).exists():
                    errores_campos['correo_inst'] = 'Este correo electrónico ya está registrado en el sistema.'
        
        # Validar CURP duplicado y formato
        if curp:
            # Validar formato (18 caracteres, patrón correcto)
            curp_upper = curp.upper().strip()
            if len(curp_upper) != 18:
                errores_campos['curp'] = 'La CURP debe tener exactamente 18 caracteres. Verifica que esté completa'
            elif not re.match(r'^[A-Z]{4}\d{6}[A-Z0-9]{6}[A-Z0-9]{2}$', curp_upper):
                errores_campos['curp'] = 'CURP inválida.'
            elif DatosPersonales.objects.filter(curp__iexact=curp).exists():
                errores_campos['curp'] = 'Esta CURP ya está registrada en el sistema. Verifica los datos'
        
        # Validar fecha de nacimiento (no puede ser futura, debe ser al menos 10 años, y año mínimo 1900)
        if fecha_nacimiento:
            try:
                fecha_obj = _parse_fecha_nacimiento(fecha_nacimiento)
                if fecha_obj:
                    hoy = timezone.now().date()
                    if fecha_obj.year < 1900:
                        errores_campos['fecha_nacimiento'] = 'El año debe ser 1900 o posterior'
                    elif fecha_obj > hoy:
                        errores_campos['fecha_nacimiento'] = 'La fecha de nacimiento no puede ser una fecha futura. Verifica el año'
                    elif (hoy - fecha_obj).days < 3650:  # Menos de 10 años
                        errores_campos['fecha_nacimiento'] = 'La fecha de nacimiento debe corresponder a una persona de al menos 10 años'
            except:
                errores_campos['fecha_nacimiento'] = 'El formato de la fecha no es válido. Usa el formato: AAAA-MM-DD (ejemplo: 2000-05-15)'
        
        # Validar género choices
        if genero:
            generos_validos = ['M', 'F', 'otro']
            if genero not in generos_validos:
                errores_campos['genero'] = 'El género seleccionado no es válido'

        foto_archivo = request.FILES.get('foto')
        if foto_archivo:
            error_foto = _validar_foto_perfil(foto_archivo)
            if error_foto:
                errores_campos['foto'] = error_foto
        
        # --- PASO 3c: HAY ERRORES → re-renderizar formulario SIN crear usuario ---
        # form_data=request.POST conserva lo que escribió el admin; errores_campos pinta mensajes rojos
        if errores_campos:
            context = {
                'proximo_id': proximo_id,
                'siguientes': siguientes,
                'año_actual': año_actual,
                'periodo_actual': _periodo_actual(),
                'carreras': Carrera.objects.order_by('nombre').all(),
                'ciclos_escolares': CicloEscolar.objects.order_by('-fecha_inicio').all(),
                'perfil': {
                    'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
                    'matricula': request.session.get('usuario_matricula', 'N/A')
                },
                'form_data': request.POST,
                'errores_campos': errores_campos,
                'primer_campo_error': _primer_campo_error_agregar_usuario(errores_campos),
                'timestamp': timezone.now().timestamp(),
            }
            return render(request, 'administrador/AgregarUsuario.html', context)
        
        # --- PASO 3d: SIN ERRORES → INSERT EN BASE DE DATOS (Create real) ---
        try:
            # transaction.atomic: si falla cualquier insert, se revierte TODO (no queda a medias)
            with transaction.atomic():
                # TABLA 1: usuarios — registro principal (matrícula y hash se aplican en .save())
                try:
                    usuario = Usuarios(
                        nombre=nombre,
                        apellido=apellido,
                        rol=rol,
                        contrasena=contrasena_temporal,
                        contrasena_temporal=True  # obligará al usuario a cambiarla al iniciar sesión
                    )
                    if foto_archivo:
                        usuario.foto = foto_archivo
                    usuario.save()  # INSERT en tabla 'usuarios'
                except ValidationError as e:
                    raise ValueError(_formatear_error_validacion(e))
                
                # TABLA 2: una sola según rol — alumnos | maestros | administrativos | administrador
                if rol == 'alumno':
                    periodo_ingreso = request.POST.get('periodo_ingreso', '').strip().upper() or _periodo_actual()
                    Alumnos.objects.create(
                        id_usuario=usuario,
                        id_carrera=Carrera.objects.get(pk=carrera_id),
                        semestre=int(semestre) if semestre else 1,
                        periodo_ingreso=periodo_ingreso,
                        estatus=estatus if estatus else 'Activo'
                    )
                
                elif rol == 'maestro':
                    Maestros.objects.create(
                        id_usuario=usuario,
                        departamento=departamento,
                        cubiculo=cubiculo if cubiculo else None,
                        grado_academico=grado_academico
                    )
                
                elif rol == 'administrativo':
                    Administrativos.objects.create(
                        id_usuario=usuario,
                        departamento=departamento,
                        puesto=puesto
                    )
                
                elif rol == 'admin':
                    ciclo_escolar = None
                    if id_ciclo_escolar:
                        try:
                            ciclo_escolar = CicloEscolar.objects.get(pk=id_ciclo_escolar)
                        except CicloEscolar.DoesNotExist:
                            pass  # ciclo opcional; si no existe, se guarda sin ciclo
                    
                    Administrador.objects.create(
                        id_usuario=usuario,
                        puesto=puesto,
                        nivel_prioridad=int(nivel_prioridad) if nivel_prioridad else 1,
                        id_ciclo_escolar=ciclo_escolar
                    )
                
                # TABLA 3: datos_personales — siempre se crea fila (campos vacíos → None)
                fecha_nacimiento_obj = _parse_fecha_nacimiento(fecha_nacimiento)
                
                try:
                    DatosPersonales.objects.create(
                        id_usuario=usuario,
                        correo_inst=correo if correo else None,
                        telefono=telefono if telefono else None,
                        curp=curp if curp else None,
                        fecha_nacimiento=fecha_nacimiento_obj,
                        genero=genero if genero else None,
                        direccion=direccion if direccion else None
                    )
                except ValidationError as e:
                    raise ValueError(_formatear_error_validacion(e))
                
                # Éxito: mensaje flash con matrícula y contraseña en texto plano (única vez visible)
                messages.success(request, _mensaje_credenciales_temporales(
                    'nuevo',
                    usuario.matricula,
                    f'{usuario.nombre} {usuario.apellido}'.strip(),
                    contrasena_temporal,
                ))
                return redirect('gestion_usuarios')  # vuelve a la lista (Read)
                
        except ValueError as e:
            # Error de validación del modelo (Usuarios.clean, DatosPersonales, etc.)
            error_msg = str(e.args[0]) if e.args else str(e)
            messages.error(request, f'Error: {error_msg}')
            return redirect('agregar_usuario')
        except Exception as e:
            # Cualquier otro fallo de BD; atomic() revierte los inserts
            messages.error(request, f'Error al crear usuario: {str(e)}')
            return redirect('agregar_usuario')
    
    # --- PASO 4: GET — primera visita o recarga; solo muestra formulario vacío ---
    context = {
        'proximo_id': proximo_id,
        'siguientes': siguientes,
        'año_actual': año_actual,
        'periodo_actual': _periodo_actual(),
        'carreras': Carrera.objects.order_by('nombre').all(),       # para <select> de carrera (alumno)
        'ciclos_escolares': CicloEscolar.objects.order_by('-fecha_inicio').all(),  # para admin
        'perfil': {
            'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
            'matricula': request.session.get('usuario_matricula', 'N/A')
        },
        'errores_campos': {},
        'primer_campo_error': '',
    }
    
    return render(request, 'administrador/AgregarUsuario.html', context)


def editar_usuario(request, usuario_id):
    """
    UPDATE (U del CRUD): modificar un usuario existente.
    URL: /administrador/usuarios/editar/<usuario_id>/  (name='editar_usuario')
    Template: administrador/EditarUsuario.html

    Flujo:
      GET  → carga usuario + datos del rol + datos personales en el formulario
      POST → valida → actualiza tablas → redirect a gestion_usuarios

    Nota: el rol NO se puede cambiar aquí (solo se muestra bloqueado en el HTML).
    """
    # --- PASO 1: SEGURIDAD ---
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')
    
    # --- PASO 2: BUSCAR USUARIO — 404 si no existe ---
    usuario = get_object_or_404(Usuarios, id_usuario=usuario_id)
    
    # --- PASO 3: CARGAR REGISTROS RELACIONADOS según rol del usuario ---
    datos_especificos = None
    datos_personales = None
    
    try:
        if usuario.rol == 'alumno':
            datos_especificos = Alumnos.objects.get(id_usuario=usuario)
        elif usuario.rol == 'maestro':
            datos_especificos = Maestros.objects.get(id_usuario=usuario)
        elif usuario.rol == 'administrativo':
            datos_especificos = Administrativos.objects.get(id_usuario=usuario)
        elif usuario.rol == 'admin':
            datos_especificos = Administrador.objects.get(id_usuario=usuario)
    except:
        pass
    
    try:
        datos_personales = DatosPersonales.objects.get(id_usuario=usuario)
    except:
        pass
    
    # --- PASO 4: POST = guardar cambios ---
    if request.method == 'POST':
        # --- PASO 4a: LEER formulario (mismos name= que en EditarUsuario.html) ---
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        correo = request.POST.get('correo_inst', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        curp = request.POST.get('curp', '').strip()
        fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()
        genero = request.POST.get('genero', '').strip()
        carrera_id = request.POST.get('carrera_id', '').strip()
        semestre = request.POST.get('semestre', '').strip()
        departamento = request.POST.get('departamento', '').strip()
        cubiculo = request.POST.get('cubiculo', '').strip()
        grado_academico = request.POST.get('grado_academico', '').strip()
        puesto = request.POST.get('puesto', '').strip()
        nivel_prioridad = request.POST.get('nivel_prioridad', '').strip()
        estatus = request.POST.get('estatus', '').strip()
        direccion = construir_direccion(request.POST)
        
        # --- PASO 4b: VALIDACIONES (igual concepto que agregar_usuario) ---
        errores_campos = {}
        
        # ============ VALIDAR CAMPOS OBLIGATORIOS ============
        if not nombre:
            errores_campos['nombre'] = 'El nombre es obligatorio'
        elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$', nombre):
            errores_campos['nombre'] = 'El nombre solo puede contener letras y espacios'
        
        if not apellido:
            errores_campos['apellido'] = 'El apellido es obligatorio'
        elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$', apellido):
            errores_campos['apellido'] = 'El apellido solo puede contener letras y espacios'
        
        # ============ VALIDAR CAMPOS ESPECÍFICOS POR ROL ============
        if usuario.rol == 'alumno' and carrera_id:
            try:
                Carrera.objects.get(pk=carrera_id)
            except Carrera.DoesNotExist:
                errores_campos['carrera_id'] = 'La carrera seleccionada no es válida. Por favor selecciona otra'
        
        if usuario.rol == 'alumno' and semestre:
            try:
                semestre_int = int(semestre)
                if semestre_int < 1 or semestre_int > 12:
                    errores_campos['semestre'] = 'El semestre debe ser un número entre 1 y 12'
            except ValueError:
                errores_campos['semestre'] = 'El semestre debe ser un número válido (ejemplo: 3)'
        
        if usuario.rol == 'alumno' and estatus:
            estatus_validos = ['Activo', 'Baja', 'Egresado']
            if estatus not in estatus_validos:
                errores_campos['estatus'] = 'El estatus seleccionado no es válido'
        
        if usuario.rol == 'maestro' and grado_academico:
            grados_validos = ['Licenciatura', 'Maestria', 'Doctorado']
            if grado_academico not in grados_validos:
                errores_campos['grado_academico'] = 'Por favor selecciona un grado académico válido'
        
        if usuario.rol == 'maestro' and not departamento:
            errores_campos['departamento'] = 'Por favor ingresa el departamento donde trabaja el maestro'
        
        if usuario.rol == 'administrativo' and not puesto:
            errores_campos['puesto'] = 'Por favor ingresa el puesto del administrativo'
        
        if usuario.rol == 'administrativo' and not departamento:
            errores_campos['departamento'] = 'Por favor ingresa el departamento del administrativo'
        
        if usuario.rol == 'admin' and puesto:
            puestos_validos = ['Director', 'Subdirector', 'Auxiliar']
            if puesto not in puestos_validos:
                errores_campos['puesto'] = 'Por favor selecciona un puesto válido de la lista'
        
        if usuario.rol == 'admin' and nivel_prioridad:
            try:
                prioridad_int = int(nivel_prioridad)
                if prioridad_int < 1 or prioridad_int > 10:
                    errores_campos['nivel_prioridad'] = 'El nivel de prioridad debe ser un número entre 1 y 10'
            except ValueError:
                errores_campos['nivel_prioridad'] = 'El nivel de prioridad debe ser un número válido (ejemplo: 5)'
        
        # ============ VALIDAR DATOS PERSONALES ============
        # Validar teléfono (solo números y 10 dígitos si se proporciona)
        if telefono:
            telefono_limpio = telefono.replace('-', '').replace(' ', '').replace('(', '').replace(')', '')
            if not telefono_limpio.isdigit():
                errores_campos['telefono'] = 'El teléfono debe contener solo números (sin guiones ni espacios)'
            elif len(telefono_limpio) != 10:
                errores_campos['telefono'] = 'El teléfono debe tener exactamente 10 dígitos'
        
        # Validar correo duplicado y formato
        if correo:
            # Validar formato de email usando EmailValidator de Django
            try:
                validate_email(correo)
            except ValidationError:
                errores_campos['correo_inst'] = 'El formato del correo no es válido.'
            else:
                # Si el formato es válido, verificar duplicados
                correos = DatosPersonales.objects.filter(correo_inst__iexact=correo).exclude(id_usuario=usuario.id_usuario)
                if correos.exists():
                    errores_campos['correo_inst'] = 'Este correo electrónico ya está registrado en el sistema.'
        
        # Validar CURP duplicado y formato
        if curp:
            # Validar formato (18 caracteres, patrón correcto)
            curp_upper = curp.upper().strip()
            if len(curp_upper) != 18:
                errores_campos['curp'] = 'La CURP debe tener exactamente 18 caracteres. Verifica que esté completa'
            elif not re.match(r'^[A-Z]{4}\d{6}[A-Z0-9]{6}[A-Z0-9]{2}$', curp_upper):
                errores_campos['curp'] = 'CURP inválida. Debe tener 18 caracteres'
            else:
                # Si el formato es válido, verificar duplicados
                curps = DatosPersonales.objects.filter(curp__iexact=curp).exclude(id_usuario=usuario.id_usuario)
                if curps.exists():
                    errores_campos['curp'] = 'Esta CURP ya está registrada en el sistema. Verifica los datos'
        
        # Validar fecha de nacimiento (no puede ser futura, debe ser al menos 10 años, y año mínimo 1900)
        if fecha_nacimiento:
            try:
                fecha_obj = _parse_fecha_nacimiento(fecha_nacimiento)
                if fecha_obj:
                    hoy = timezone.now().date()
                    if fecha_obj.year < 1900:
                        errores_campos['fecha_nacimiento'] = 'El año debe ser 1900 o posterior'
                    elif fecha_obj > hoy:
                        errores_campos['fecha_nacimiento'] = 'La fecha de nacimiento no puede ser una fecha futura. Verifica el año'
                    elif (hoy - fecha_obj).days < 3650:  # Menos de 10 años
                        errores_campos['fecha_nacimiento'] = 'La fecha de nacimiento debe corresponder a una persona de al menos 10 años'
            except:
                errores_campos['fecha_nacimiento'] = 'El formato de la fecha no es válido. Usa el formato: AAAA-MM-DD (ejemplo: 2000-05-15)'
        
        # Validar género choices
        if genero:
            generos_validos = ['M', 'F', 'otro']
            if genero not in generos_validos:
                errores_campos['genero'] = 'El género seleccionado no es válido'
        
        # --- PASO 4c: ERRORES → re-renderizar EditarUsuario.html sin guardar ---
        if errores_campos:
            context = {
                'usuario': usuario,
                'datos_especificos': datos_especificos,
                'datos_personales': datos_personales,
                'carreras': Carrera.objects.order_by('nombre').all(),
                'direccion_data': desglosar_direccion(datos_personales.direccion if datos_personales else None),
                'alumno_data': {
                    'id_carrera_id': datos_especificos.id_carrera_id if usuario.rol == 'alumno' and datos_especificos else '',
                    'semestre': datos_especificos.semestre if usuario.rol == 'alumno' and datos_especificos else '',
                    'periodo_ingreso': datos_especificos.periodo_ingreso if usuario.rol == 'alumno' and datos_especificos else '',
                    'estatus': datos_especificos.estatus if usuario.rol == 'alumno' and datos_especificos else 'Activo',
                },
                'maestro_data': {
                    'departamento': datos_especificos.departamento if usuario.rol == 'maestro' and datos_especificos else '',
                    'cubiculo': datos_especificos.cubiculo if usuario.rol == 'maestro' and datos_especificos else '',
                    'grado_academico': datos_especificos.grado_academico if usuario.rol == 'maestro' and datos_especificos else '',
                },
                'administrativo_data': {
                    'departamento': datos_especificos.departamento if usuario.rol == 'administrativo' and datos_especificos else '',
                    'puesto': datos_especificos.puesto if usuario.rol == 'administrativo' and datos_especificos else '',
                },
                'admin_data': {
                    'puesto': datos_especificos.puesto if usuario.rol == 'admin' and datos_especificos else '',
                    'nivel_prioridad': datos_especificos.nivel_prioridad if usuario.rol == 'admin' and datos_especificos else 1,
                },
                'datos_personales_data': {
                    'correo_inst': correo,
                    'telefono': telefono,
                    'curp': curp,
                    'fecha_nacimiento': fecha_nacimiento,
                    'genero': genero,
                    'direccion': direccion,
                },
                'form_data': request.POST,
                'errores_campos': errores_campos,
                'primer_campo_error': _primer_campo_error_editar_usuario(errores_campos),
                'timestamp': timezone.now().timestamp(),
                'perfil': {
                    'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
                    'matricula': request.session.get('usuario_matricula', 'N/A')
                }
            }
            return render(request, 'administrador/EditarUsuario.html', _anexar_retorno_lista_usuarios(request, context))
        
        # --- PASO 4d: SIN ERRORES → UPDATE en BD (transaction.atomic) ---
        try:
            with transaction.atomic():
                # TABLA usuarios: actualizar nombre, apellido, foto, contraseña opcional
                usuario.nombre = nombre
                usuario.apellido = apellido

                # request.FILES: foto de perfil (todos los roles)
                foto_archivo = request.FILES.get('foto')
                if foto_archivo:
                    error_foto = _validar_foto_perfil(foto_archivo)
                    if error_foto:
                        errores_campos['foto'] = error_foto
                        context = {
                            'usuario': usuario,
                            'datos_especificos': datos_especificos,
                            'datos_personales': datos_personales,
                            'carreras': Carrera.objects.order_by('nombre').all(),
                            'direccion_data': desglosar_direccion(datos_personales.direccion if datos_personales else None),
                            'alumno_data': {
                                'id_carrera_id': datos_especificos.id_carrera_id if usuario.rol == 'alumno' and datos_especificos else '',
                                'semestre': datos_especificos.semestre if usuario.rol == 'alumno' and datos_especificos else '',
                                'periodo_ingreso': datos_especificos.periodo_ingreso if usuario.rol == 'alumno' and datos_especificos else '',
                                'estatus': datos_especificos.estatus if usuario.rol == 'alumno' and datos_especificos else 'Activo',
                            },
                            'maestro_data': {
                                'departamento': datos_especificos.departamento if usuario.rol == 'maestro' and datos_especificos else '',
                                'cubiculo': datos_especificos.cubiculo if usuario.rol == 'maestro' and datos_especificos else '',
                                'grado_academico': datos_especificos.grado_academico if usuario.rol == 'maestro' and datos_especificos else '',
                            },
                            'administrativo_data': {
                                'departamento': datos_especificos.departamento if usuario.rol == 'administrativo' and datos_especificos else '',
                                'puesto': datos_especificos.puesto if usuario.rol == 'administrativo' and datos_especificos else '',
                            },
                            'admin_data': {
                                'puesto': datos_especificos.puesto if usuario.rol == 'admin' and datos_especificos else '',
                                'nivel_prioridad': datos_especificos.nivel_prioridad if usuario.rol == 'admin' and datos_especificos else 1,
                            },
                            'datos_personales_data': {
                                'correo_inst': correo,
                                'telefono': telefono,
                                'curp': curp,
                                'fecha_nacimiento': fecha_nacimiento,
                                'genero': genero,
                                'direccion': direccion,
                            },
                            'form_data': request.POST,
                            'errores_campos': errores_campos,
                            'primer_campo_error': _primer_campo_error_editar_usuario(errores_campos),
                            'timestamp': timezone.now().timestamp(),
                            'perfil': {
                                'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
                                'matricula': request.session.get('usuario_matricula', 'N/A')
                            }
                        }
                        return render(request, 'administrador/EditarUsuario.html', _anexar_retorno_lista_usuarios(request, context))

                    if usuario.foto:
                        usuario.foto.delete(save=False)
                    usuario.foto = foto_archivo
                elif request.POST.get('quitar_foto') == '1':
                    _quitar_foto_perfil(usuario)

                # Contraseña nueva opcional; si va vacía, no se cambia
                nueva_contrasena = request.POST.get('contrasena', '').strip()
                if nueva_contrasena:
                    usuario.contrasena = nueva_contrasena

                try:
                    usuario.save()
                except ValidationError as e:
                    raise ValueError(_formatear_error_validacion(e))
                
                # TABLA del rol: .save() actualiza fila existente (no .create())
                if usuario.rol == 'alumno' and datos_especificos:
                    carrera_id = request.POST.get('carrera_id', '').strip()
                    if carrera_id:
                        datos_especificos.id_carrera = Carrera.objects.get(pk=carrera_id)
                    semestre = request.POST.get('semestre', '').strip()
                    if semestre:
                        datos_especificos.semestre = int(semestre)
                    datos_especificos.estatus = request.POST.get('estatus', '').strip()
                    datos_especificos.save()
                
                elif usuario.rol == 'maestro' and datos_especificos:
                    datos_especificos.departamento = request.POST.get('departamento', '').strip()
                    datos_especificos.cubiculo = request.POST.get('cubiculo', '').strip() or None
                    datos_especificos.grado_academico = request.POST.get('grado_academico', '').strip()
                    datos_especificos.save()
                
                elif usuario.rol == 'administrativo' and datos_especificos:
                    datos_especificos.departamento = request.POST.get('departamento', '').strip()
                    datos_especificos.puesto = request.POST.get('puesto', '').strip()
                    datos_especificos.save()
                
                elif usuario.rol == 'admin' and datos_especificos:
                    datos_especificos.puesto = request.POST.get('puesto', '').strip()
                    nivel_prioridad = request.POST.get('nivel_prioridad', '').strip()
                    if nivel_prioridad:
                        datos_especificos.nivel_prioridad = int(nivel_prioridad)
                    datos_especificos.save()
                
                # TABLA datos_personales: update si existe; create si no había fila
                correo = request.POST.get('correo_inst', '').strip()
                telefono = request.POST.get('telefono', '').strip()
                curp = request.POST.get('curp', '').strip()
                fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()
                genero = request.POST.get('genero', '').strip()
                direccion = construir_direccion(request.POST)
                
                fecha_nacimiento_obj = _parse_fecha_nacimiento(fecha_nacimiento)

                error_unicos = _validar_datos_personales_unicos(correo, curp, usuario.id_usuario)
                if error_unicos:
                    raise ValueError(error_unicos)

                if datos_personales:
                    datos_personales.correo_inst = correo if correo else None
                    datos_personales.telefono = telefono if telefono else None
                    datos_personales.curp = curp if curp else None
                    datos_personales.fecha_nacimiento = fecha_nacimiento_obj
                    datos_personales.genero = genero if genero else None
                    datos_personales.direccion = direccion if direccion else None
                    try:
                        datos_personales.save()
                    except ValidationError as e:
                        raise ValueError(_formatear_error_validacion(e))
                elif correo or telefono or curp or direccion or fecha_nacimiento_obj or genero:
                    try:
                        DatosPersonales.objects.create(
                            id_usuario=usuario,
                            correo_inst=correo if correo else None,
                            telefono=telefono if telefono else None,
                            curp=curp if curp else None,
                            fecha_nacimiento=fecha_nacimiento_obj,
                            genero=genero if genero else None,
                            direccion=direccion if direccion else None
                        )
                    except ValidationError as e:
                        raise ValueError(_formatear_error_validacion(e))
                
                messages.success(request, f'¡Cambios guardados! Se actualizó la información de {usuario.nombre}')
                return _redirect_gestion_usuarios(request)
                
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')
            return redirect(
                reverse('editar_usuario', args=[usuario_id]) + _query_lista_usuarios(_filtros_lista_usuarios(request))
            )
    
    # --- PASO 5: GET — mostrar formulario prellenado con datos actuales ---
    context = {
        'usuario': usuario,
        'datos_especificos': datos_especificos,
        'datos_personales': datos_personales,
        'carreras': Carrera.objects.order_by('nombre').all(),
        'direccion_data': desglosar_direccion(datos_personales.direccion if datos_personales else None),
        'alumno_data': {
            'id_carrera_id': datos_especificos.id_carrera_id if usuario.rol == 'alumno' and datos_especificos else '',
            'semestre': datos_especificos.semestre if usuario.rol == 'alumno' and datos_especificos else '',
            'periodo_ingreso': datos_especificos.periodo_ingreso if usuario.rol == 'alumno' and datos_especificos else '',
            'estatus': datos_especificos.estatus if usuario.rol == 'alumno' and datos_especificos else 'Activo',
        },
        'maestro_data': {
            'departamento': datos_especificos.departamento if usuario.rol == 'maestro' and datos_especificos else '',
            'cubiculo': datos_especificos.cubiculo if usuario.rol == 'maestro' and datos_especificos else '',
            'grado_academico': datos_especificos.grado_academico if usuario.rol == 'maestro' and datos_especificos else '',
        },
        'administrativo_data': {
            'departamento': datos_especificos.departamento if usuario.rol == 'administrativo' and datos_especificos else '',
            'puesto': datos_especificos.puesto if usuario.rol == 'administrativo' and datos_especificos else '',
        },
        'admin_data': {
            'puesto': datos_especificos.puesto if usuario.rol == 'admin' and datos_especificos else '',
            'nivel_prioridad': datos_especificos.nivel_prioridad if usuario.rol == 'admin' and datos_especificos else 1,
        },
        'datos_personales_data': {
            'correo_inst': datos_personales.correo_inst if datos_personales else '',
            'telefono': datos_personales.telefono if datos_personales else '',
            'curp': datos_personales.curp if datos_personales else '',
            'fecha_nacimiento': datos_personales.fecha_nacimiento.isoformat() if datos_personales and datos_personales.fecha_nacimiento else '',
            'genero': datos_personales.genero if datos_personales else '',
            'direccion': datos_personales.direccion if datos_personales else '',
        },
        'perfil': {
            'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
            'matricula': request.session.get('usuario_matricula', 'N/A')
        },
        'errores_campos': {},
        'primer_campo_error': '',
    }
    
    return render(request, 'administrador/EditarUsuario.html', _anexar_retorno_lista_usuarios(request, context))


def eliminar_usuario(request, usuario_id):
    """
    DELETE (D del CRUD): borrar un usuario.
    URL: /administrador/usuarios/eliminar/<usuario_id>/  (name='eliminar_usuario')

    Flujo:
      GET  → confirmar_eliminar.html (poco usado; la lista usa AJAX)
      POST → usuario.delete() en cascada → redirect o JsonResponse

    Regla: no se puede eliminar al último admin del sistema.
    """
    # --- PASO 1: SEGURIDAD — respuesta JSON si es petición AJAX sin permiso ---
    if not sesion_roles_permitidas(request, ('admin',)):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos'}, status=403)
        return redirect('selector_rol')
    
    # --- PASO 2: BUSCAR usuario a eliminar ---
    usuario = get_object_or_404(Usuarios, id_usuario=usuario_id)
    
    # --- PASO 3: POST = ejecutar borrado ---
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Regla de negocio: debe quedar al menos 1 administrador
                if usuario.rol == 'admin':
                    total_admins = Usuarios.objects.filter(rol='admin').count()
                    if total_admins <= 1:
                        error_msg = 'No puedes eliminar al último administrador del sistema'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': error_msg}, status=400)
                        messages.error(request, error_msg)
                        return redirect('gestion_usuarios')
                
                # DELETE: OneToOne CASCADE borra alumnos/maestros/datos_personales ligados
                usuario_nombre = f"{usuario.nombre} {usuario.apellido}"
                admin_nombre = request.session.get('usuario_nombre', 'Administrador')
                logger.info(f'Usuario eliminado: {usuario_nombre} (ID: {usuario_id}) por {admin_nombre}')
                usuario.delete()
                
                # Éxito: JSON para AJAX (GestionUsuarios.html) o redirect clásico
                success_msg = f'Usuario {usuario_nombre} eliminado exitosamente'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': success_msg})
                messages.success(request, success_msg)
                return redirect('gestion_usuarios')
                
        except Exception as e:
            logger.error(f'Error al eliminar usuario {usuario_id}: {str(e)}')
            error_msg = f'Error al eliminar usuario: {str(e)}'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
            messages.error(request, error_msg)
            return redirect('gestion_usuarios')
    
    # --- PASO 4: GET — pantalla de confirmación (alternativa al modal AJAX) ---
    context = {
        'usuario': usuario,
        'perfil': {
            'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
            'matricula': request.session.get('usuario_matricula', 'N/A')
        }
    }
    
    return render(request, 'administrador/confirmar_eliminar.html', context)


def gestion_seguridad(request):
    """Vista para gestionar la seguridad del sistema"""
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')
    
    # Estadísticas de seguridad
    total_usuarios = Usuarios.objects.count()
    usuarios_activos = Usuarios.objects.filter(ultimo_acceso__isnull=False).count()
    usuarios_por_rol = Usuarios.objects.values('rol').count()
    
    # Últimos accesos
    ultimos_accesos = Usuarios.objects.filter(
        ultimo_acceso__isnull=False
    ).order_by('-ultimo_acceso')[:10]
    
    # Cuentas bloqueadas
    cuentas_bloqueadas = Usuarios.objects.filter(
        cuenta_bloqueada=True
    ).order_by('-fecha_bloqueo')[:10]
    
    # Logs de calificaciones recientes
    logs_calificaciones = LogCalificacion.objects.select_related(
        'id_usuario_modifico'
    ).order_by('-fecha_modificacion')[:10]
    
    context = {
        'total_usuarios': total_usuarios,
        'usuarios_activos': usuarios_activos,
        'usuarios_por_rol': usuarios_por_rol,
        'logs_acceso': ultimos_accesos,
        'cuentas_bloqueadas': cuentas_bloqueadas,
        'logs_calificaciones': logs_calificaciones,
        'ultimo_respaldo': 'N/A',  # Se puede implementar después
        'perfil': {
            'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
            'matricula': request.session.get('usuario_matricula', 'N/A')
        }
    }
    
    return render(request, 'administrador/GestionSeguridad.html', context)


def desbloquear_cuenta(request, usuario_id):
    """Desbloquea una cuenta de usuario manualmente"""
    if not sesion_roles_permitidas(request, ('admin',)):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'})
        return redirect('selector_rol')

    if request.method != 'POST':
        return redirect('gestion_seguridad')

    try:
        usuario = Usuarios.objects.get(id_usuario=usuario_id)

        if usuario.cuenta_bloqueada:
            usuario.cuenta_bloqueada = False
            usuario.intentos_fallidos_login = 0
            usuario.fecha_bloqueo = None
            usuario.save()

            admin_nombre = request.session.get('usuario_nombre', 'Administrador')
            logger.info(f'Cuenta desbloqueada: {usuario.nombre} {usuario.apellido} (ID: {usuario_id}) por {admin_nombre}')

            mensaje = f'✓ Cuenta de {usuario.nombre} {usuario.apellido} desbloqueada exitosamente.'

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': mensaje})

            messages.success(request, mensaje)
        else:
            mensaje = 'La cuenta no está bloqueada.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': mensaje})
            messages.warning(request, mensaje)

    except Usuarios.DoesNotExist:
        logger.warning(f'Intento de desbloquear usuario inexistente: {usuario_id}')
        mensaje = 'Usuario no encontrado.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': mensaje})
        messages.error(request, mensaje)
    except Exception as e:
        logger.error(f'Error al desbloquear cuenta {usuario_id}: {str(e)}')
        mensaje = f'Error al desbloquear cuenta: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': mensaje})
        messages.error(request, mensaje)

    return redirect('gestion_seguridad')


@transaction.atomic
def restablecer_contrasena(request, usuario_id):
    """
    UPDATE parcial: genera nueva contraseña temporal (no es CRUD puro, va con gestión).
    URL: /administrador/usuarios/restablecer/<usuario_id>/  (name='restablecer_contrasena')
    Solo POST — se dispara desde modal en GestionUsuarios.html
    """
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')

    if request.method != 'POST':
        return _redirect_gestion_usuarios(request)

    usuario = get_object_or_404(Usuarios, id_usuario=usuario_id)

    try:
        nueva_contrasena = generar_contrasena_temporal()

        # UPDATE en tabla usuarios: nueva clave + flag contrasena_temporal=True
        usuario.contrasena = nueva_contrasena
        usuario.contrasena_temporal = True
        usuario.save()

        messages.success(request, _mensaje_credenciales_temporales(
            'restablecer',
            usuario.matricula,
            f'{usuario.nombre} {usuario.apellido}'.strip(),
            nueva_contrasena,
        ))
    except Exception as e:
        messages.error(request, f"Error al restablecer contraseña: {str(e)}")

    return _redirect_gestion_usuarios(request)


def _resolver_ruta_pg_dump() -> str | None:
    """Ubica pg_dump en PATH o en instalaciones comunes de PostgreSQL (Windows)."""
    configurado = getattr(settings, 'PG_DUMP_PATH', None) or os.environ.get('PG_DUMP_PATH')
    if configurado and os.path.isfile(configurado):
        return configurado

    encontrado = shutil.which('pg_dump')
    if encontrado:
        return encontrado

    if os.name == 'nt':
        bases = [
            os.environ.get('ProgramFiles', r'C:\Program Files'),
            os.environ.get('ProgramFiles(x86)', r'C:\Program Files (x86)'),
        ]
        for base in bases:
            pg_root = os.path.join(base, 'PostgreSQL')
            if not os.path.isdir(pg_root):
                continue
            versiones = sorted(os.listdir(pg_root), reverse=True)
            for version in versiones:
                candidato = os.path.join(pg_root, version, 'bin', 'pg_dump.exe')
                if os.path.isfile(candidato):
                    return candidato

    return None


def respaldo_bdd(request):
    """Vista para gestionar respaldos de la base de datos"""
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')
    
    from django.conf import settings
    import glob
    
    # Obtener lista de respaldos existentes
    backup_dir = settings.BACKUP_DIR
    backup_files = []
    
    if os.path.exists(backup_dir):
        for file_path in glob.glob(os.path.join(backup_dir, '*.sql')):
            try:
                stat = os.stat(file_path)
                size_mb = stat.st_size / (1024 * 1024)
                backup_files.append({
                    'nombre': os.path.basename(file_path),
                    'ruta': file_path,
                    'tamaño': f"{size_mb:.2f} MB",
                    'fecha': datetime.fromtimestamp(stat.st_mtime).strftime('%d/%m/%Y %H:%M:%S')
                })
            except:
                pass
    
    # Ordenar por fecha (más reciente primero)
    backup_files.sort(key=lambda x: x['fecha'], reverse=True)
    
    # Generar nombre para próximo respaldo
    fecha_respaldo = datetime.now().strftime('%Y%m%d_%H%M%S')
    ruta_respaldo = os.path.join(backup_dir, f"respaldo_schooltrack_{fecha_respaldo}.sql")
    
    context = {
        'perfil': {
            'nombre_completo': request.session.get('usuario_nombre', 'Administrador'),
            'matricula': request.session.get('usuario_matricula', 'N/A')
        },
        'fecha_respaldo': fecha_respaldo,
        'ruta_respaldo': ruta_respaldo,
        'backup_files': backup_files,
        'total_backups': len(backup_files)
    }
    
    return render(request, 'administrador/RespaldoBDD.html', context)


def ejecutar_respaldo(request):
    """Vista para ejecutar un respaldo de la base de datos"""
    if not sesion_roles_permitidas(request, ('admin',)):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'No tienes permisos para realizar esta acción'})
        return redirect('selector_rol')
    
    if request.method == 'POST':
        try:
            from django.conf import settings
            import subprocess
            
            # Obtener configuración de base de datos
            db_config = settings.DATABASES['default']
            backup_dir = settings.BACKUP_DIR
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = os.path.join(backup_dir, f"respaldo_schooltrack_{timestamp}.sql")
            
            # Verificar si el archivo ya existe (confirmación de sobrescritura)
            if os.path.exists(backup_file):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False, 
                        'message': f'El archivo {os.path.basename(backup_file)} ya existe. Elimínelo manualmente o espere al próximo respaldo.'
                    })
                messages.error(request, f'El archivo ya existe. Elimínelo manualmente o espere al próximo respaldo.')
                return redirect('respaldo_bdd')
            
            pg_dump_bin = _resolver_ruta_pg_dump()
            if not pg_dump_bin:
                error_msg = (
                    'No se encontró pg_dump en el sistema. '
                    'Instala las herramientas de PostgreSQL o define PG_DUMP_PATH en tu archivo .env.'
                )
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
                return redirect('respaldo_bdd')

            # Construir comando pg_dump
            pg_dump_cmd = [
                pg_dump_bin,
                f'--host={db_config["HOST"]}',
                f'--port={db_config["PORT"]}',
                f'--username={db_config["USER"]}',
                f'--dbname={db_config["NAME"]}',
                '--no-password',
                '--format=plain',
                '--no-owner',
                '--no-acl',
                f'--file={backup_file}'
            ]
            
            # Establecer variable de entorno para contraseña
            env = os.environ.copy()
            env['PGPASSWORD'] = db_config['PASSWORD']
            
            # Ejecutar pg_dump
            result = subprocess.run(pg_dump_cmd, env=env, capture_output=True, text=True)
            
            if result.returncode != 0:
                error_msg = result.stderr or 'Error desconocido al ejecutar pg_dump'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': f'Error al ejecutar pg_dump: {error_msg}'})
                messages.error(request, f'Error al ejecutar pg_dump: {error_msg}')
                return redirect('respaldo_bdd')
            
            # Obtener tamaño del archivo
            file_size = os.path.getsize(backup_file)
            size_mb = file_size / (1024 * 1024)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'message': f'Respaldo realizado exitosamente. Tamaño: {size_mb:.2f} MB',
                    'filename': os.path.basename(backup_file),
                    'size_mb': f'{size_mb:.2f} MB'
                })
            
            messages.success(request, f'Respaldo realizado exitosamente. Tamaño: {size_mb:.2f} MB')
            return redirect('respaldo_bdd')
            
        except Exception as e:
            logger.error(f"Error al realizar respaldo: {str(e)}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': f'Error al realizar respaldo: {str(e)}'})
            
            messages.error(request, f'Error al realizar respaldo: {str(e)}')
            return redirect('respaldo_bdd')
    
    return redirect('respaldo_bdd')


def descargar_respaldo_especifico(request, filename):
    """Vista para descargar un respaldo específico"""
    if not sesion_roles_permitidas(request, ('admin',)):
        return redirect('selector_rol')
    
    try:
        from django.conf import settings
        import mimetypes
        
        backup_dir = settings.BACKUP_DIR
        file_path = os.path.join(backup_dir, filename)
        
        # Verificar que el archivo existe y está dentro del directorio de respaldos
        if not os.path.exists(file_path) or not os.path.abspath(file_path).startswith(os.path.abspath(backup_dir)):
            messages.error(request, 'El archivo de respaldo no existe o no es válido')
            return redirect('respaldo_bdd')
        
        # Determinar el tipo MIME
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type is None:
            mime_type = 'application/octet-stream'
        
        # Leer el archivo
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        # Crear respuesta HTTP
        response = HttpResponse(file_content, content_type=mime_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = str(len(file_content))
        
        return response
        
    except Exception as e:
        logger.error(f"Error al descargar respaldo: {str(e)}")
        messages.error(request, f'Error al descargar respaldo: {str(e)}')
        return redirect('respaldo_bdd')
